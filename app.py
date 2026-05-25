import logging
"""
app.py – Interfaz gráfica principal (Tkinter)
Control de Stock v1.0
"""
import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import database as db


# ──────────────────────────────────────────────
#  Paleta de colores
# ──────────────────────────────────────────────
BG        = "#1e1e2e"
BG2       = "#2a2a3e"
BG3       = "#313145"
ACCENT    = "#7c6af7"
ACCENT2   = "#a89ef9"
TEXT      = "#e0e0f0"
TEXT_DIM  = "#888899"
SUCCESS   = "#4caf82"
DANGER    = "#e06c75"
WARNING   = "#e5c07b"
ROW_ALT   = "#252535"


# ──────────────────────────────────────────────
#  Helpers de estilo
# ──────────────────────────────────────────────

def styled_btn(parent, text, command, color=ACCENT, fg=TEXT, **kw):
    return tk.Button(
        parent, text=text, command=command,
        bg=color, fg=fg, relief="flat",
        padx=12, pady=6,
        font=("Segoe UI", 9, "bold"),
        activebackground=ACCENT2, activeforeground=TEXT,
        cursor="hand2", **kw
    )


def lbl(parent, text, size=10, bold=False, color=TEXT, **kw):
    weight = "bold" if bold else "normal"
    return tk.Label(parent, text=text, bg=BG2, fg=color,
                    font=("Segoe UI", size, weight), **kw)


def entry(parent, textvariable=None, width=20, **kw):
    e = tk.Entry(parent, textvariable=textvariable, width=width,
                 bg=BG3, fg=TEXT, insertbackground=TEXT,
                 relief="flat", font=("Segoe UI", 10),
                 highlightthickness=1, highlightcolor=ACCENT,
                 highlightbackground=BG3, **kw)
    return e

def fmt_qty(val, por_peso=0):
    """Format quantity to hide .0 if int and add 'Kg' if por_peso is true"""
    try:
        fval = float(val)
        res = str(int(fval)) if fval.is_integer() else str(fval)
        return f"{res} Kg" if por_peso else res
    except:
        return str(val)


# ──────────────────────────────────────────────
#  Ventana de alta / edición de producto
# ──────────────────────────────────────────────

class ProductoDialog(tk.Toplevel):
    def __init__(self, parent, producto=None):
        super().__init__(parent)
        self.title("Nuevo Producto" if producto is None else "Editar Producto")
        self.configure(bg=BG2)
        self.bind("<Escape>", lambda e: self.destroy())
        self.resizable(False, False)
        self.grab_set()          # modal
        self.resultado = None
        self._producto = producto

        # Variables
        self.v_codigo    = tk.StringVar(value=producto["codigo"]    if producto else "")
        self.v_nombre    = tk.StringVar(value=producto["nombre"]    if producto else "")
        self.v_categoria = tk.StringVar(value=producto["categoria"] if producto else "")
        self.v_marca     = tk.StringVar(value=producto.get("marca", "") if producto else "")
        self.v_stock     = tk.StringVar(value=fmt_qty(producto["stock"], producto.get("por_peso")).replace(" Kg", "")   if producto else "0")
        self.v_minimo    = tk.StringVar(value=fmt_qty(producto["minimo"], producto.get("por_peso")).replace(" Kg", "")  if producto else "0")
        self.v_precio    = tk.StringVar(value=str(producto["precio"])  if producto else "0.0")
        self.v_precio_costo = tk.StringVar(value=str(producto.get("precio_costo", 0.0)) if producto else "0.0")
        self.v_por_peso  = tk.BooleanVar(value=bool(producto.get("por_peso", 0)) if producto else False)

        frame = tk.Frame(self, bg=BG2, padx=24, pady=20)
        frame.pack(fill="both", expand=True)

        fields = [
            ("Código *",    self.v_codigo),
            ("Nombre *",    self.v_nombre),
            ("Categoría",   self.v_categoria),
            ("Marca",       self.v_marca),
            ("Stock inicial", self.v_stock),
            ("Stock mínimo",  self.v_minimo),
            ("Precio Costo",  self.v_precio_costo),
            ("Precio Venta",  self.v_precio),
        ]

        self.lbl_stock = None
        self.lbl_precio = None

        # Configurar campos desplegables (Combobox)
        self.cmb_cat = ttk.Combobox(frame, textvariable=self.v_categoria, values=db.get_categorias(), width=26)
        self.cmb_marca = ttk.Combobox(frame, textvariable=self.v_marca, values=db.get_marcas(), width=26)

        for i, (label_txt, var) in enumerate(fields):
            lbl = tk.Label(frame, text=label_txt, bg=BG2, fg=TEXT_DIM, font=("Segoe UI", 9))
            lbl.grid(row=i, column=0, sticky="w", pady=5, padx=(0,12))
            
            if "Stock inicial" in label_txt: self.lbl_stock = lbl
            if "Precio Venta" in label_txt: self.lbl_precio = lbl
            if "Precio Costo" in label_txt: self.lbl_costo = lbl
            
            if label_txt == "Categoría":
                self.cmb_cat.grid(row=i, column=1, sticky="ew", pady=5)
            elif label_txt == "Marca":
                self.cmb_marca.grid(row=i, column=1, sticky="ew", pady=5)
            elif label_txt == "Código *":
                code_frame = tk.Frame(frame, bg=BG2)
                code_frame.grid(row=i, column=1, sticky="ew", pady=5)
                entry(code_frame, textvariable=var, width=16).pack(side="left", fill="x", expand=True)
                styled_btn(code_frame, "⚡", self._autogenerar_codigo, color="#2563eb", width=3).pack(side="left", padx=(5, 0))
            else:
                entry(frame, textvariable=var, width=28).grid(row=i, column=1, sticky="ew", pady=5)

        tk.Checkbutton(frame, text="Se vende por peso (Permite decimales)", variable=self.v_por_peso,
                       bg=BG2, fg=TEXT_DIM, selectcolor=BG, activebackground=BG2,
                       font=("Segoe UI", 9)).grid(row=len(fields), column=0, columnspan=2, sticky="w", pady=8)

        def _update_labels(*args):
            suf = " (Kg)" if self.v_por_peso.get() else ""
            if self.lbl_stock: self.lbl_stock.config(text=f"Stock inicial{suf}")
            if self.lbl_precio: self.lbl_precio.config(text=f"Precio Venta{suf}")
            if hasattr(self, 'lbl_costo') and self.lbl_costo: self.lbl_costo.config(text=f"Precio Costo{suf}")
            
        self.v_por_peso.trace_add("write", _update_labels)
        _update_labels()

        btn_frame = tk.Frame(frame, bg=BG2)
        btn_frame.grid(row=len(fields)+1, column=0, columnspan=2, pady=(16, 0))

        styled_btn(btn_frame, "💾 Guardar", self._guardar, color=ACCENT).pack(side="left", padx=6)
        styled_btn(btn_frame, "✖ Cancelar", self.destroy, color=BG3).pack(side="left", padx=6)

        self.wait_window()

    def _autogenerar_codigo(self):
        try:
            conn = db.get_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT codigo FROM productos")
            codigos = [row[0] for row in cursor.fetchall() if row[0] and row[0].isdigit()]
            siguiente = max(map(int, codigos)) + 1 if codigos else 1
            self.v_codigo.set(str(siguiente))
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo generar el código: {e}", parent=self)

    def _guardar(self):
        codigo    = self.v_codigo.get().strip()
        nombre    = self.v_nombre.get().strip()
        categoria = self.v_categoria.get().strip()
        marca     = self.v_marca.get().strip()

        if not codigo or not nombre:
            messagebox.showwarning("Campo requerido", "Código y Nombre son obligatorios.", parent=self)
            return

        try:
            stock_str = self.v_stock.get().replace(" Kg", "")
            stock  = float(stock_str)
            minimo_str = self.v_minimo.get().replace(" Kg", "")
            minimo = float(minimo_str)
            precio_costo = float(self.v_precio_costo.get())
            precio = float(self.v_precio.get())
            por_peso = 1 if self.v_por_peso.get() else 0
        except ValueError:
            messagebox.showerror("Datos inválidos", "Stock, mínimo y precios deben ser numéricos.", parent=self)
            return

        self.resultado = (codigo, nombre, categoria, marca, stock, minimo, precio, por_peso, precio_costo)
        self.destroy()

class EdicionMasivaDialog(tk.Toplevel):
    def __init__(self, parent, count):
        super().__init__(parent)
        self.title("Edición Masiva de Productos")
        self.configure(bg=BG2)
        self.bind("<Escape>", lambda e: self.destroy())
        self.resizable(False, False)
        self.grab_set()
        
        self.resultado = None  # { "categoria": val, "marca": val, "por_peso": val }
        
        # Variables de control (Checkbox)
        self.chk_cat_var = tk.BooleanVar(value=False)
        self.chk_marca_var = tk.BooleanVar(value=False)
        self.chk_peso_var = tk.BooleanVar(value=False)
        
        # Variables de valor
        self.v_categoria = tk.StringVar()
        self.v_marca = tk.StringVar()
        self.v_por_peso = tk.BooleanVar(value=False)
        
        frame = tk.Frame(self, bg=BG2, padx=24, pady=20)
        frame.pack(fill="both", expand=True)

        tk.Label(frame, text=f"Se editarán {count} productos seleccionados.", bg=BG2, fg=ACCENT2, font=("Segoe UI", 10, "bold")).grid(row=0, column=0, columnspan=2, pady=(0, 15), sticky="w")
        
        # --- Categoría ---
        chk_cat = tk.Checkbutton(frame, text="Modificar Categoría:", variable=self.chk_cat_var,
                                 bg=BG2, fg=TEXT, selectcolor=BG, activebackground=BG2, font=("Segoe UI", 9))
        chk_cat.grid(row=1, column=0, sticky="w", pady=5)
        
        self.cmb_cat = ttk.Combobox(frame, textvariable=self.v_categoria, values=db.get_categorias(), width=26, state="disabled")
        self.cmb_cat.grid(row=1, column=1, sticky="w", padx=(10,0))
        
        # --- Marca ---
        chk_marca = tk.Checkbutton(frame, text="Modificar Marca:", variable=self.chk_marca_var,
                                   bg=BG2, fg=TEXT, selectcolor=BG, activebackground=BG2, font=("Segoe UI", 9))
        chk_marca.grid(row=2, column=0, sticky="w", pady=5)
        
        self.cmb_marca = ttk.Combobox(frame, textvariable=self.v_marca, values=db.get_marcas(), width=26, state="disabled")
        self.cmb_marca.grid(row=2, column=1, sticky="w", padx=(10,0))
        
        # --- Por Peso ---
        chk_peso = tk.Checkbutton(frame, text="Modificar 'Venta por Peso':", variable=self.chk_peso_var,
                                  bg=BG2, fg=TEXT, selectcolor=BG, activebackground=BG2, font=("Segoe UI", 9))
        chk_peso.grid(row=3, column=0, sticky="w", pady=5)
        
        self.check_peso_val = tk.Checkbutton(frame, text="Sí, se vende por peso", variable=self.v_por_peso,
                                             bg=BG2, fg=TEXT_DIM, selectcolor=BG, activebackground=BG2, font=("Segoe UI", 9), state="disabled")
        self.check_peso_val.grid(row=3, column=1, sticky="w", padx=(10,0))

        # Trace para habilitar/deshabilitar
        def toggle_cat(*args):
            self.cmb_cat.config(state="normal" if self.chk_cat_var.get() else "disabled")
        def toggle_marca(*args):
            self.cmb_marca.config(state="normal" if self.chk_marca_var.get() else "disabled")
        def toggle_peso(*args):
            self.check_peso_val.config(state="normal" if self.chk_peso_var.get() else "disabled")

        self.chk_cat_var.trace_add("write", toggle_cat)
        self.chk_marca_var.trace_add("write", toggle_marca)
        self.chk_peso_var.trace_add("write", toggle_peso)

        btn_frame = tk.Frame(frame, bg=BG2)
        btn_frame.grid(row=4, column=0, columnspan=2, pady=(20, 0))

        styled_btn(btn_frame, "💾 Aplicar a Todos", self._guardar, color=ACCENT).pack(side="left", padx=6)
        styled_btn(btn_frame, "✖ Cancelar", self.destroy, color=BG3).pack(side="left", padx=6)

        self.wait_window()

    def _guardar(self):
        cambios = {}
        if self.chk_cat_var.get():
            cambios["categoria"] = self.v_categoria.get().strip()
        if self.chk_marca_var.get():
            cambios["marca"] = self.v_marca.get().strip()
        if self.chk_peso_var.get():
            cambios["por_peso"] = 1 if self.v_por_peso.get() else 0

        if not cambios:
            messagebox.showwarning("Atención", "No has seleccionado ningún campo para modificar.", parent=self)
            return

        resp = messagebox.askyesno("Confirmar", f"Se modificarán {len(cambios)} campos a todos los productos seleccionados.\n\n¿Estás seguro?", parent=self)
        if resp:
            self.resultado = cambios
            self.destroy()


# ──────────────────────────────────────────────
#  Ventana de movimiento (entrada / salida)
# ──────────────────────────────────────────────

class MovimientoDialog(tk.Toplevel):
    def __init__(self, parent, producto, tipo):
        super().__init__(parent)
        self.title(f"{'Entrada' if tipo=='entrada' else 'Salida'} de Stock")
        self.configure(bg=BG2)
        self.bind("<Escape>", lambda e: self.destroy())
        self.resizable(False, False)
        self.grab_set()
        self.resultado = None
        self.producto = producto
        self._precio_base = float(producto.get("precio", 0) or 0)
        self._actualizando = False  # bandera para evitar recursión en trace

        color_tipo = SUCCESS if tipo == "entrada" else DANGER

        frame = tk.Frame(self, bg=BG2, padx=24, pady=20)
        frame.pack(fill="both", expand=True)

        tk.Label(frame, text=f"{'▲ ENTRADA' if tipo=='entrada' else '▼ SALIDA'}",
                 bg=BG2, fg=color_tipo, font=("Segoe UI", 12, "bold")).grid(row=0, column=0, columnspan=2, pady=(0,10))

        info = f"{producto['nombre']}  |  Stock actual: {fmt_qty(producto['stock'], producto.get('por_peso'))}"
        tk.Label(frame, text=info, bg=BG2, fg=TEXT, font=("Segoe UI", 10)).grid(row=1, column=0, columnspan=2, pady=(0,4))

        # Precio base
        precio_txt = f"Precio base: ${self._precio_base:,.2f}" if self._precio_base else "Precio base: no definido"
        tk.Label(frame, text=precio_txt, bg=BG2, fg=TEXT_DIM, font=("Segoe UI", 9, "italic")).grid(row=2, column=0, columnspan=2, pady=(0,10))

        self.v_cantidad = tk.StringVar(value="1")
        self.v_nota     = tk.StringVar()

        tk.Label(frame, text="Cantidad *", bg=BG2, fg=TEXT_DIM, font=("Segoe UI", 9)).grid(row=3, column=0, sticky="w", pady=5, padx=(0,12))
        entry(frame, textvariable=self.v_cantidad, width=14).grid(row=3, column=1, sticky="ew", pady=5)

        tk.Label(frame, text="Nota", bg=BG2, fg=TEXT_DIM, font=("Segoe UI", 9)).grid(row=4, column=0, sticky="w", pady=5, padx=(0,12))
        entry(frame, textvariable=self.v_nota, width=28).grid(row=4, column=1, sticky="ew", pady=5)

        btn_frame = tk.Frame(frame, bg=BG2)
        btn_frame.grid(row=5, column=0, columnspan=2, pady=(16, 0))

        styled_btn(btn_frame, "\u2714 Confirmar", lambda: self._confirmar(producto, tipo),
                   color=color_tipo).pack(side="left", padx=6)
        styled_btn(btn_frame, "\u2716 Cancelar", self.destroy, color=BG3).pack(side="left", padx=6)

        # ── Panel precio/monto (solo en salida, y solo si hay precio definido) ──
        if tipo == "salida" and self._precio_base > 0:
            tk.Frame(self, bg=BG3, height=1).pack(fill="x", padx=16, pady=(8, 0))
            self._build_precio_panel()

        self.wait_window()

    def _build_precio_panel(self):
        """Panel bidireccional: editar Unidades actualiza Monto y viceversa."""
        pf = tk.Frame(self, bg=BG2, padx=24, pady=14)
        pf.pack(fill="x")

        tk.Label(pf, text="\U0001f4b0  Referencia de precio", bg=BG2, fg=ACCENT2,
                 font=("Segoe UI", 9, "bold")).grid(row=0, column=0, columnspan=3, sticky="w", pady=(0, 8))

        tk.Label(pf, text="Unidades", bg=BG2, fg=TEXT_DIM, font=("Segoe UI", 8)).grid(row=1, column=0, padx=(0,6))
        tk.Label(pf, text="Monto ($)", bg=BG2, fg=TEXT_DIM, font=("Segoe UI", 8)).grid(row=1, column=2, padx=(6,0))

        self.v_ref_unid  = tk.StringVar()
        self.v_ref_monto = tk.StringVar()

        e_unid  = entry(pf, textvariable=self.v_ref_unid,  width=10)
        e_monto = entry(pf, textvariable=self.v_ref_monto, width=12)
        e_unid.grid(row=2, column=0, padx=(0,4))
        tk.Label(pf, text="\u2194", bg=BG2, fg=TEXT_DIM, font=("Segoe UI", 14)).grid(row=2, column=1)
        e_monto.grid(row=2, column=2, padx=(4,0))

        # Etiqueta informativa dinámica
        self.lbl_ref_info = tk.Label(pf, text="", bg=BG2, fg=SUCCESS, font=("Segoe UI", 9, "bold"))
        self.lbl_ref_info.grid(row=3, column=0, columnspan=3, pady=(6,0))

        # Botón para pasar unidades al campo cantidad principal
        styled_btn(pf, "\u2191 Usar unidades como cantidad", lambda: self._usar_unidades_ref(self.producto),
                   color=BG3).grid(row=4, column=0, columnspan=3, pady=(8, 0))

        # Traces bidireccionales
        self.v_ref_unid.trace_add("write",  lambda *_: self._sync_desde_unidades())
        self.v_ref_monto.trace_add("write", lambda *_: self._sync_desde_monto())

    def _sync_desde_unidades(self):
        if self._actualizando:
            return
        self._actualizando = True
        try:
            u = float(self.v_ref_unid.get().replace(",", "."))
            monto = u * self._precio_base
            self.v_ref_monto.set(f"{monto:,.2f}")
            self.lbl_ref_info.config(
                text=f"{u:,.0f} unid. × ${self._precio_base:,.2f} = ${monto:,.2f}",
                fg=SUCCESS)
        except (ValueError, AttributeError):
            try:
                self.lbl_ref_info.config(text="", fg=SUCCESS)
            except Exception as e: logging.exception("Error silencioso capturado:")
        finally:
            self._actualizando = False

    def _sync_desde_monto(self):
        if self._actualizando:
            return
        self._actualizando = True
        try:
            m = float(self.v_ref_monto.get().replace(",", "."))
            if self._precio_base > 0:
                unidades = m / self._precio_base
                self.v_ref_unid.set(f"{unidades:,.2f}")
                self.lbl_ref_info.config(
                    text=f"${m:,.2f} \u00f7 ${self._precio_base:,.2f} = {unidades:,.2f} unid.",
                    fg=ACCENT2)
        except (ValueError, AttributeError):
            try:
                self.lbl_ref_info.config(text="", fg=SUCCESS)
            except Exception as e: logging.exception("Error silencioso capturado:")
        finally:
            self._actualizando = False

    def _usar_unidades_ref(self, producto):
        try:
            val = float(self.v_ref_unid.get().replace(",", "."))
            if not producto.get("por_peso", 0):
                val = int(val)
            if val > 0:
                self.v_cantidad.set(str(val))
        except (ValueError, AttributeError):
            pass

    def _confirmar(self, producto, tipo):
        try:
            val_str = self.v_cantidad.get().replace(",", ".")
            if producto.get("por_peso", 0):
                cantidad = float(val_str)
            else:
                cantidad = int(float(val_str))  # Permitir si tipean 5.0, limpiarlo a 5
            
            if cantidad <= 0:
                raise ValueError("La cantidad debe ser mayor a cero.")
        except ValueError:
            msg = "Ingrese un número válido." if producto.get("por_peso", 0) else "Ingrese un número entero positivo."
            messagebox.showerror("Cantidad inválida", msg, parent=self)
            return

        nota = self.v_nota.get().strip()
        try:
            precio_total = (cantidad * producto.get("precio", 0.0)) if tipo == "salida" else None
            db.registrar_movimiento(producto["id"], tipo, cantidad, nota, forzar=False, precio_total=precio_total)
            self.resultado = True
            self.destroy()
        except ValueError as e:
            if "insuficiente" in str(e).lower():
                stock_actual = producto["stock"]
                resp = messagebox.askyesno(
                    "\u26a0\ufe0f Stock insuficiente",
                    f"Stock actual: {stock_actual}  |  Cantidad solicitada: {cantidad}\n\n"
                    f"\u00bfDesea registrar la salida de todas formas?\n"
                    f"(Se marcar\u00e1 como excepci\u00f3n en el historial)",
                    icon="warning", parent=self
                )
                if resp:
                    nota_exc = (nota + " [EXCEPCI\u00d3N: stock insuficiente]").strip()
                    precio_total = (cantidad * producto.get("precio", 0.0)) if tipo == "salida" else None
                    db.registrar_movimiento(producto["id"], tipo, cantidad, nota_exc, forzar=True, precio_total=precio_total)
                    self.resultado = True
                    self.destroy()
            else:
                messagebox.showerror("Error", str(e), parent=self)


class ModificarItemDialog(tk.Toplevel):
    def __init__(self, parent, producto, cant_inicial, nota_inicial, desc_val=0, desc_tipo="%", desc_acum=True):
        super().__init__(parent)
        self.title("Modificar cantidad")
        self.configure(bg=BG2)
        self.resizable(False, False)
        self.grab_set()
        
        self.resultado = None
        self.producto = producto

        frame = tk.Frame(self, bg=BG2, padx=20, pady=16)
        frame.pack(fill="both", expand=True)
        
        prod_name = str(producto.get("nombre", ""))
        tk.Label(frame, text=prod_name, bg=BG2, fg=ACCENT2, font=("Segoe UI", 10, "bold")).grid(row=0, column=0, columnspan=2, pady=(0, 10))

        self.v_cantidad   = tk.StringVar(value=fmt_qty(cant_inicial, False))
        self.v_nota       = tk.StringVar(value=str(nota_inicial))
        
        # Nuevas variables de Descuento Individual
        self.v_desc_val   = tk.StringVar(value=str(desc_val))
        self.v_desc_tipo  = tk.StringVar(value=desc_tipo)
        self.v_desc_acum  = tk.BooleanVar(value=desc_acum)

        self._is_peso = bool(self.producto.get("por_peso", 0))
        self.v_monto = tk.StringVar()
        precio_uni = float(self.producto.get("precio", 0.0))

        tk.Label(frame, text=f"Precio unitario: ${precio_uni:,.2f}{' / Kg' if self._is_peso else ''}", bg=BG2, fg=TEXT, font=("Segoe UI", 9, "italic")).grid(row=1, column=0, columnspan=2, pady=(0, 10))
        
        lbl_cant = "Peso (Kg) *" if self._is_peso else "Cantidad *"
        tk.Label(frame, text=lbl_cant, bg=BG2, fg=TEXT_DIM, font=("Segoe UI", 9)).grid(row=2, column=0, sticky="w", pady=5, padx=(0,12))
        self.e_cant = entry(frame, textvariable=self.v_cantidad, width=14)
        self.e_cant.grid(row=2, column=1, sticky="ew", pady=5)
        
        tk.Label(frame, text="Monto Base ($)", bg=BG2, fg=TEXT_DIM, font=("Segoe UI", 9)).grid(row=3, column=0, sticky="w", pady=5, padx=(0,12))
        e_monto = entry(frame, textvariable=self.v_monto, width=14)
        e_monto.grid(row=3, column=1, sticky="ew", pady=5)

        tk.Label(frame, text="Nota", bg=BG2, fg=TEXT_DIM, font=("Segoe UI", 9)).grid(row=4, column=0, sticky="w", pady=5, padx=(0,12))
        entry(frame, textvariable=self.v_nota, width=28).grid(row=4, column=1, sticky="ew", pady=5)

        self._syncing = False
        
        def _calc_monto(*args):
            if self._syncing: return
            try:
                val = float(self.v_cantidad.get().replace(",", "."))
                self._syncing = True
                self.v_monto.set(f"{val * precio_uni:,.2f}")
                self._syncing = False
            except ValueError:
                self._syncing = True
                self.v_monto.set("")
                self._syncing = False

        def _calc_peso(*args):
            if self._syncing: return
            try:
                if precio_uni <= 0: return
                val = float(self.v_monto.get().replace(",", "."))
                self._syncing = True
                if self._is_peso:
                    self.v_cantidad.set(f"{val / precio_uni:.3f}")
                else:
                    self.v_cantidad.set(f"{int(val / precio_uni)}")
                self._syncing = False
            except ValueError:
                self._syncing = True
                self.v_cantidad.set("")
                self._syncing = False
                
        self.v_cantidad.trace_add("write", _calc_monto)
        self.v_monto.trace_add("write", _calc_peso)
        _calc_monto() # trigger first compute
            
        # --- Descuentos Individuales ---
        desc_frame = tk.Frame(frame, bg=BG2)
        desc_frame.grid(row=5, column=0, columnspan=2, pady=(10, 5), sticky="w")

        tk.Label(desc_frame, text="Descuento:", bg=BG2, fg=TEXT_DIM, font=("Segoe UI", 9)).pack(side="left", padx=(0, 5))
        self.e_desc = entry(desc_frame, textvariable=self.v_desc_val, width=8)
        self.e_desc.pack(side="left")
        cmb_tipo = ttk.Combobox(desc_frame, textvariable=self.v_desc_tipo, values=["%", "$"], width=3, state="readonly")
        cmb_tipo.pack(side="left", padx=5)
        
        tk.Checkbutton(desc_frame, text="Acumulable c/ Global", variable=self.v_desc_acum, bg=BG2, fg=TEXT, selectcolor=BG3, activebackground=BG2, activeforeground=TEXT, font=("Segoe UI", 9)).pack(side="left", padx=(10, 0))

        btn_frame = tk.Frame(frame, bg=BG2)
        btn_frame.grid(row=6, column=0, columnspan=2, pady=(16, 0))

        styled_btn(btn_frame, "\u2714 Confirmar", self._confirmar, color=SUCCESS).pack(side="left", padx=6)
        styled_btn(btn_frame, "\u2716 Cancelar", self.destroy, color=BG3).pack(side="left", padx=6)

        self.bind("<Return>", lambda e: self._confirmar() if type(self.focus_get()) != ttk.Combobox else None)
        self.bind("<Escape>", lambda e: self.destroy())
        
        # --- Navegación interna con flechas ---
        widgets_nav = [self.e_cant, e_monto, self.e_desc, cmb_tipo]
        
        def _nav_up_down(event, delta):
            current = self.focus_get()
            if current in widgets_nav:
                idx = widgets_nav.index(current)
                next_idx = (idx + delta) % len(widgets_nav)
                widgets_nav[next_idx].focus_set()
            return "break"
            
        def _nav_left_right(event, delta):
            current = self.focus_get()
            if current in [self.e_desc, cmb_tipo]:
                if current == self.e_desc and delta == 1:
                    cmb_tipo.focus_set()
                elif current == cmb_tipo and delta == -1:
                    self.e_desc.focus_set()
                return "break"
                
        self.bind("<Down>", lambda e: _nav_up_down(e, 1))
        self.bind("<Up>", lambda e: _nav_up_down(e, -1))
        self.bind("<Left>", lambda e: _nav_left_right(e, -1))
        self.bind("<Right>", lambda e: _nav_left_right(e, 1))
        
        cmb_tipo.bind("<Return>", lambda e: cmb_tipo.event_generate('<Down>'))

        self.e_cant.focus_set()
        self.wait_window()
        
    def _confirmar(self):
        try:
            val_str = self.v_cantidad.get().replace(",", ".")
            if self._is_peso:
                cant = float(val_str)
            else:
                cant = int(float(val_str))
                
            
            # Sanitizar descuento
            try:
                desc_v = float(self.v_desc_val.get().replace(",", "."))
                if desc_v < 0: raise ValueError
            except:
                desc_v = 0.0
                
            self.resultado = {
                "cant": cant, 
                "nota": self.v_nota.get().strip(),
                "desc_val": desc_v,
                "desc_tipo": self.v_desc_tipo.get(),
                "desc_acum": self.v_desc_acum.get()
            }
            self.destroy()
        except ValueError:
            msg = "Valor numérico mayor a 0 requerido." if self._is_peso else "Cantidad entera mayor a 0 requerida."
            messagebox.showerror("Inválido", msg, parent=self)

class EvolucionPreciosDialog(tk.Toplevel):
    def __init__(self, parent, cambios):
        super().__init__(parent)
        self.title("Evolución de Precios")
        self.configure(bg=BG2)
        self.bind("<Escape>", lambda e: self.destroy())
        self.geometry("500x350")
        self.resizable(True, True)
        self.grab_set()

        frame = tk.Frame(self, bg=BG2, padx=16, pady=16)
        frame.pack(fill="both", expand=True)

        tk.Label(frame, text="Resumen de Modificaciones", bg=BG2, fg=WARNING, font=("Segoe UI", 12, "bold")).pack(pady=(0, 10))

        # Crear Treeview
        columns = ("producto", "precio_ant", "precio_nuev")
        self.tree = ttk.Treeview(frame, columns=columns, show="headings", height=10)
        self.tree.bind("<Button-1>", lambda e: "break" if self.tree.identify_region(e.x, e.y) == "separator" else None)
        
        self.tree.heading("producto", text="Producto")
        self.tree.heading("precio_ant", text="Precio Anterior")
        self.tree.heading("precio_nuev", text="Precio Nuevo")
        
        self.tree.column("producto", width=250, anchor="w")
        self.tree.column("precio_ant", width=100, anchor="e")
        self.tree.column("precio_nuev", width=100, anchor="e")

        # Scrollbar
        scrollbar = ttk.Scrollbar(frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        self.tree.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Insertar datos
        for idx, (nombre, ant, nuev) in enumerate(cambios):
            tag = "alt" if idx % 2 == 1 else ""
            self.tree.insert("", "end", values=(nombre, f"${ant:,.2f}", f"${nuev:,.2f}"), tags=(tag,))
            
        self.tree.tag_configure("alt", background=BG1)

        btn_frame = tk.Frame(self, bg=BG2, pady=10)
        btn_frame.pack(fill="x", side="bottom")

        styled_btn(btn_frame, "✔ Aceptar", self.destroy, color=SUCCESS).pack(pady=5)
        
        self.wait_window()

class AumentoMasivoDialog(tk.Toplevel):
    def __init__(self, parent, productos_seleccionados=None):
        super().__init__(parent)
        self.title("Ajuste Masivo de Precios")
        self.configure(bg=BG2)
        self.bind("<Escape>", lambda e: self.destroy())
        self.resizable(False, False)
        self.grab_set()

        self.resultado = False
        self.parent_app = parent
        self.productos_seleccionados = productos_seleccionados or []

        frame = tk.Frame(self, bg=BG2, padx=20, pady=16)
        frame.pack(fill="both", expand=True)

        tk.Label(frame, text="Ajustar Precios", bg=BG2, fg=WARNING, font=("Segoe UI", 12, "bold")).grid(row=0, column=0, columnspan=2, pady=(0, 10))
        
        self.lbl_sel_text = tk.Label(frame, text="", bg=BG2, fg=ACCENT, font=("Segoe UI", 9, "bold"))
        self.lbl_sel_text.grid(row=1, column=0, columnspan=2, pady=(0, 10))
        self._refresh_lbl_count()

        # Checkboxes y Comboboxes para Filtros
        self.v_check_cat = tk.BooleanVar(value=False)
        self.v_check_marca = tk.BooleanVar(value=False)
        self.v_cat = tk.StringVar(value="Cualquiera")
        self.v_marca = tk.StringVar(value="Cualquiera")

        chk_cat = tk.Checkbutton(frame, text="Filtrar por Categoría", variable=self.v_check_cat, command=self._toggle_filtros, bg=BG2, fg=TEXT_DIM, selectcolor=BG3, activebackground=BG2, activeforeground=TEXT, font=("Segoe UI", 9))
        chk_cat.grid(row=2, column=0, sticky="w", pady=5)
        self.cmb_cat = ttk.Combobox(frame, textvariable=self.v_cat, state="disabled", width=25)
        self.cmb_cat["values"] = ["Cualquiera"] + db.get_categorias()
        self.cmb_cat.grid(row=2, column=1, sticky="ew", pady=5, padx=(10, 0))

        chk_marca = tk.Checkbutton(frame, text="Filtrar por Marca", variable=self.v_check_marca, command=self._toggle_filtros, bg=BG2, fg=TEXT_DIM, selectcolor=BG3, activebackground=BG2, activeforeground=TEXT, font=("Segoe UI", 9))
        chk_marca.grid(row=3, column=0, sticky="w", pady=5)
        self.cmb_marca = ttk.Combobox(frame, textvariable=self.v_marca, state="disabled", width=25)
        self.cmb_marca["values"] = ["Cualquiera"] + db.get_marcas()
        self.cmb_marca.grid(row=3, column=1, sticky="ew", pady=5, padx=(10, 0))

        # Tipo de Aumento
        self.v_tipo = tk.StringVar(value="porcentaje")
        
        radio_frame = tk.Frame(frame, bg=BG2)
        radio_frame.grid(row=4, column=0, columnspan=2, sticky="w", pady=(15, 5))
        
        tk.Label(radio_frame, text="Método:", bg=BG2, fg=TEXT, font=("Segoe UI", 9)).pack(side="left", padx=(0, 10))
        tk.Radiobutton(radio_frame, text="Porcentaje (%)", variable=self.v_tipo, value="porcentaje", bg=BG2, fg=TEXT, selectcolor=BG3, activebackground=BG2, font=("Segoe UI", 9)).pack(side="left", padx=5)
        tk.Radiobutton(radio_frame, text="Monto Fijo ($)", variable=self.v_tipo, value="fijo", bg=BG2, fg=TEXT, selectcolor=BG3, activebackground=BG2, font=("Segoe UI", 9)).pack(side="left", padx=5)

        # Monto/Porcentaje
        tk.Label(frame, text="Valor a incrementar *", bg=BG2, fg=TEXT_DIM, font=("Segoe UI", 9)).grid(row=5, column=0, sticky="w", pady=5)
        self.v_valor = tk.StringVar()
        entry_calc = entry(frame, textvariable=self.v_valor, width=15)
        entry_calc.grid(row=5, column=1, sticky="w", pady=5, padx=(10, 0))

        # --- Previsualización (Solo si se seleccionaron manualmente) ---
        self.tree_prev = None
        if self.productos_seleccionados:
            prev_frame = tk.Frame(frame, bg=BG)
            prev_frame.grid(row=6, column=0, columnspan=2, pady=(15, 0), sticky="nsew")
            tk.Label(prev_frame, text="Previsualización de Impacto [Suprimir/Backspace para Remover]", bg=BG, fg=TEXT_DIM, font=("Segoe UI", 9, "bold")).pack(anchor="w", padx=5, pady=2)
            
            cols = ("prod", "precio_ant", "precio_new")
            self.tree_prev = ttk.Treeview(prev_frame, columns=cols, show="headings", height=8, selectmode="extended")
            self.tree_prev.bind("<Button-1>", lambda e: "break" if self.tree_prev.identify_region(e.x, e.y) == "separator" else None)
            self.tree_prev.heading("prod", text="Producto")
            self.tree_prev.heading("precio_ant", text="Precio Actual")
            self.tree_prev.heading("precio_new", text="Nuevo Precio")
            self.tree_prev.column("prod", width=400, anchor="w")
            self.tree_prev.column("precio_ant", width=120, anchor="center")
            self.tree_prev.column("precio_new", width=120, anchor="center")
            
            sb_prev = ttk.Scrollbar(prev_frame, orient="vertical", command=self.tree_prev.yview)
            self.tree_prev.configure(yscrollcommand=sb_prev.set)
            self.tree_prev.pack(side="left", fill="both", expand=True, padx=(5, 0), pady=(0, 5))
            sb_prev.pack(side="right", fill="y", padx=(0, 5), pady=(0, 5))
            
            self._fill_tree_prev()
                
            self.v_valor.trace_add("write", self._actualizar_preview)
            self.v_tipo.trace_add("write", self._actualizar_preview)
            
            def _remove_selected(event):
                if not self.tree_prev: return
                sel = self.tree_prev.selection()
                if not sel: return
                for iid in sel:
                    prod_id = int(iid)
                    self.productos_seleccionados = [p for p in self.productos_seleccionados if p["id"] != prod_id]
                self._fill_tree_prev()
                self._actualizar_preview()
                self._refresh_lbl_count()
            
            self.tree_prev.bind("<BackSpace>", _remove_selected)
            self.tree_prev.bind("<Delete>", _remove_selected)

        btn_frame = tk.Frame(frame, bg=BG2)
        btn_frame.grid(row=7, column=0, columnspan=2, pady=(20, 0))

        styled_btn(btn_frame, "✔ Aplicar", self._aplicar, color=WARNING).pack(side="left", padx=6)
        styled_btn(btn_frame, "✖ Cancelar", self.destroy, color=BG3).pack(side="left", padx=6)

        # Centrar la ventana y desplazarla más arriba para evitar la barra de tareas
        self.update_idletasks()
        w = self.winfo_reqwidth()
        h = self.winfo_reqheight()
        x = (self.winfo_screenwidth() // 2) - (w // 2)
        # Desplazamos la ventana 80px hacia arriba desde el centro absoluto
        y = max(0, (self.winfo_screenheight() // 2) - (h // 2) - 80)
        self.geometry(f"+{x}+{y}")

        self.wait_window()
        
    def _refresh_lbl_count(self):
        total_sel = len(self.productos_seleccionados)
        text = f"{total_sel} Producto(s) seleccionado(s) manualmente." if total_sel > 0 else "0 Productos seleccionados manualmente."
        self.lbl_sel_text.config(text=text)

    def _fill_tree_prev(self):
        if not self.tree_prev: return
        self.tree_prev.delete(*self.tree_prev.get_children())
        for p in self.productos_seleccionados:
            self.tree_prev.insert("", "end", iid=str(p["id"]), values=(p["nombre"], f"${p['precio']:,.2f}", "-"))

    def _actualizar_preview(self, *args):
        if not self.tree_prev: return
        
        try:
            val_str = self.v_valor.get().replace(",", ".")
            valor = float(val_str)
        except ValueError:
            valor = 0.0
            
        tipo = self.v_tipo.get()
        # Dictionary for faster access since iid is the product id
        p_dict = {str(p["id"]): p for p in self.productos_seleccionados}
        
        for item_id in self.tree_prev.get_children():
            p = p_dict.get(item_id)
            if not p: continue
            
            precio_base = p["precio"]
            if valor == 0:
                nuevo_txt = "-"
            else:
                if tipo == "porcentaje":
                    nuevo = precio_base * (1 + valor / 100.0)
                else:
                    nuevo = precio_base + valor
                nuevo_txt = f"${nuevo:,.2f}"
            
            self.tree_prev.item(item_id, values=(p["nombre"], f"${precio_base:,.2f}", nuevo_txt))
            
    def _toggle_filtros(self):
        if self.v_check_cat.get():
            self.cmb_cat.config(state="readonly")
        else:
            self.cmb_cat.config(state="disabled")
            self.v_cat.set("Cualquiera")
            
        if self.v_check_marca.get():
            self.cmb_marca.config(state="readonly")
        else:
            self.cmb_marca.config(state="disabled")
            self.v_marca.set("Cualquiera")

    def _aplicar(self):
        try:
            val_str = self.v_valor.get().replace(",", ".")
            valor = float(val_str)
            if valor == 0: raise ValueError
        except ValueError:
            messagebox.showerror("Error", "Ingrese un importe válido distinto de 0.", parent=self)
            return

        cat = self.v_cat.get() if self.v_check_cat.get() else "Cualquiera"
        marca = self.v_marca.get() if self.v_check_marca.get() else "Cualquiera"
        tipo = self.v_tipo.get()
        
        is_global = not self.productos_seleccionados and cat == "Cualquiera" and marca == "Cualquiera"
        if is_global:
            # Check if explicitly authorized by checking both boxes and selecting "Cualquiera"
            authorized = self.v_check_cat.get() and self.v_check_marca.get()
            if not authorized:
                messagebox.showerror("Ajuste Global Bloqueado", "Para aplicar un ajuste a TODOS los productos, debe activar las casillas de filtro para Categoría y Marca, dejándolas en 'Cualquiera'.", parent=self)
                return

        simbolo = "%" if tipo == "porcentaje" else "$"
        post_simbolo = "" if tipo == "porcentaje" else " fijos"
        
        tipo_accion = "un aumento" if valor >= 0 else "una baja"
        valor_abs = abs(valor)
        msg = f"¿Está seguro de aplicar {tipo_accion} de {simbolo}{valor_abs}{post_simbolo} "
        if self.productos_seleccionados:
            msg += f"\na los {len(self.productos_seleccionados)} productos seleccionados manualmente?"
        elif cat != "Cualquiera" and marca != "Cualquiera":
            msg += f"\na los productos de categoría '{cat}' y marca '{marca}'?"
        elif cat != "Cualquiera":
            msg += f"\na los productos de la categoría '{cat}'?"
        elif marca != "Cualquiera":
            msg += f"\na los productos de la marca '{marca}'?"
        else:
            msg += f"\na TODOS los productos globalmente?"

        resp = messagebox.askyesno("Confirmar Ajuste", msg, icon="warning", parent=self)
        if resp:
            ids_to_update = [p["id"] for p in self.productos_seleccionados]
            cambios = db.aplicar_aumento_masivo(valor, tipo, cat, marca, ids_to_update)
            self.resultado = True
            self.destroy()
            
            if cambios:
                EvolucionPreciosDialog(self.parent_app, cambios)
            else:
                messagebox.showinfo("Resultado", "No se encontraron productos que coincidan con el criterio para aumentar.")



class SalidaMultipleDialog(tk.Toplevel):
    def __init__(self, parent, productos_seleccionados):
        super().__init__(parent)
        self.title("Registro de Salidas Múltiples (Carrito)")
        self.geometry("1024x768")
        self.minsize(800, 600)
        self.state("zoomed")
        self.configure(bg=BG)
        self.bind("<Escape>", lambda e: self.destroy())
        self.grab_set()

        self.resultado = False
        self.parent_app = parent
        
        # State
        self.items = {}
        for p in productos_seleccionados:
            self.items[p["id"]] = {
                "prod": p, 
                "cant": p.get("_init_cant", 1), 
                "nota": "", 
                "desc_val": 0, 
                "desc_tipo": "%", 
                "desc_acum": True
            }
            
        self.v_desc_global_val = tk.StringVar(value="0")
        self.v_desc_global_tipo = tk.StringVar(value="%")
        self.v_paga_con = tk.StringVar(value="")
        self._total_final_vuelto = 0.0
            
        self._build_ui()
        self._refresh_list()
        self.grab_set()
        self.focus_set()
        self.wait_window()

    def _build_ui(self):
        header = tk.Frame(self, bg=BG2, pady=10)
        header.pack(fill="x")
        tk.Label(header, text="🛒 Salida Múltiple", bg=BG2, fg=DANGER, font=("Segoe UI", 14, "bold")).pack()
        tk.Label(header, text="Haga doble clic en un ítem para modificar su cantidad o nota.", bg=BG2, fg=TEXT_DIM, font=("Segoe UI", 9)).pack()

        # Treeview
        tree_frame = tk.Frame(self, bg=BG)
        tree_frame.pack(fill="both", expand=True, padx=16, pady=12)

        cols = ("id", "nombre", "stock", "precio", "cant", "desc", "bruto", "subtotal", "nota")
        headers = ("ID", "Producto", "Stock", "Precio Unit.", "Cant.", "Desc.", "Precio S/Desc.", "Subtotal", "Nota")
        widths = (0, 200, 70, 90, 60, 60, 100, 90, 160)

        self.tree = ttk.Treeview(tree_frame, columns=cols, show="headings", selectmode="browse")
        self.tree.bind("<Button-1>", lambda e: "break" if self.tree.identify_region(e.x, e.y) == "separator" else None)
        for col, hdr, w in zip(cols, headers, widths):
            self.tree.heading(col, text=hdr)
            self.tree.column(col, width=w, anchor="center" if col not in ("nombre", "nota") else "w")
            
        self.tree.column("id", stretch=tk.NO, width=0) # Hide ID

        sb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)
        
        self.tree.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")
        
        self.tree.bind("<Double-1>", self._on_double_click)
        self.tree.bind("<Return>", self._on_key_enter)
        self.tree.bind("<BackSpace>", lambda e: self._quitar_item())
        self.tree.bind("<Delete>", lambda e: self._quitar_item())
        
        # Enlazar en el Toplevel y explícitamente en el árbol
        def finalize_sale(e=None):
            self._registrar_todos()
            return "break"
        self.bind("<F12>", finalize_sale)
        self.tree.bind("<F12>", finalize_sale)
        self.bind("<Escape>", lambda e: self.destroy())

        def _tree_select_first(event):
            if not isinstance(event.widget, ttk.Treeview):
                items = self.tree.get_children()
                if items:
                    self.tree.focus(items[0])
                    self.tree.selection_set(items[0])
                    self.tree.focus_set()
                    return "break"
        self.bind("<Down>", _tree_select_first)
        self.bind("<Up>", _tree_select_first)
        
        # Bottom panel
        bot_frame = tk.Frame(self, bg=BG, pady=10)
        bot_frame.pack(fill="x", padx=16, pady=(0, 10))
        
        # Global discount controls
        desc_global_frame = tk.Frame(bot_frame, bg=BG)
        desc_global_frame.pack(side="left", padx=10)
        
        tk.Label(desc_global_frame, text="Desc. Global:", bg=BG, fg=TEXT_DIM, font=("Segoe UI", 9)).pack(side="left", padx=(0,5))
        e_glob_desc = entry(desc_global_frame, textvariable=self.v_desc_global_val, width=6)
        e_glob_desc.pack(side="left")
        e_glob_desc.bind("<KeyRelease>", lambda e: self._refresh_list())
        
        cmb_glob_tipo = ttk.Combobox(desc_global_frame, textvariable=self.v_desc_global_tipo, values=["%", "$"], width=3, state="readonly")
        cmb_glob_tipo.pack(side="left", padx=5)
        cmb_glob_tipo.bind("<<ComboboxSelected>>", lambda e: self._refresh_list())
        
        # Cambio / Vuelto
        vuelto_frame = tk.Frame(bot_frame, bg=BG)
        vuelto_frame.pack(side="left", padx=20)
        
        tk.Label(vuelto_frame, text="Paga con: $", bg=BG, fg=TEXT_DIM, font=("Segoe UI", 9)).pack(side="left")
        self.e_paga_con = entry(vuelto_frame, textvariable=self.v_paga_con, width=8)
        self.e_paga_con.pack(side="left")
        self.v_paga_con.trace_add("write", lambda *a: self._calc_vuelto())
        
        self.lbl_vuelto = tk.Label(vuelto_frame, text="Vuelto: -", bg=BG, fg=SUCCESS, font=("Segoe UI", 10, "bold"))
        self.lbl_vuelto.pack(side="left", padx=(10, 0))
        
        self.lbl_total = tk.Label(bot_frame, text="Total: $0.00", bg=BG, fg=TEXT, font=("Segoe UI", 16, "bold"))
        self.lbl_total.pack(side="left", padx=15)
        
        styled_btn(bot_frame, "✖ Quitar selec.", self._quitar_item, color=BG3).pack(side="left", padx=5)
        
        styled_btn(bot_frame, "✔ Confirmar (F12)", self._registrar_todos, color=DANGER).pack(side="right")
        styled_btn(bot_frame, "Cancelar", self.destroy, color=BG3).pack(side="right", padx=10)

    def _refresh_list(self):
        self.tree.delete(*self.tree.get_children())
        total_previo = 0.0
        
        # Parse Global Discount
        try:
            glob_val = float(self.v_desc_global_val.get().replace(",", "."))
        except:
            glob_val = 0.0
        glob_tipo = self.v_desc_global_tipo.get()
        
        # Primera pasada: Calcular total bruto para sacar % proporcional de un descuento global en $
        for pid, data in self.items.items():
            precio = float(data["prod"].get("precio", 0) or 0)
            total_previo += precio * float(data["cant"])
            
        porc_global_efectivo = 0.0
        if glob_val > 0:
            if glob_tipo == "%":
                porc_global_efectivo = glob_val / 100.0
            else: # es $
                if total_previo > 0:
                    porc_global_efectivo = glob_val / total_previo
                    
        total_monto = 0.0
        
        for pid, data in self.items.items():
            p = data["prod"]
            cant = data["cant"]
            nota = data["nota"]
            precio_val = p.get("precio", 0)
            if not precio_val: precio_val = 0
            precio = float(precio_val) # type: ignore
            subtotal_bruto = precio * float(cant)
            
            # --- Matemáticas de Descuentos ---
            # 1. Descuento individual
            desc_ind_val = float(data.get("desc_val", 0))
            desc_ind_tipo = data.get("desc_tipo", "%")
            acumulable = data.get("desc_acum", True)
            
            subtotal_item = subtotal_bruto
            str_desc = ""
            if desc_ind_val > 0:
                if desc_ind_tipo == "%":
                    subtotal_item -= subtotal_bruto * (desc_ind_val / 100.0)
                    str_desc = f"-{desc_ind_val}%"
                else:
                    subtotal_item -= desc_ind_val
                    str_desc = f"-${desc_ind_val:,.2f}"
            
            # 2. Descuento global
            if porc_global_efectivo > 0:
                if desc_ind_val == 0 or acumulable:
                    # Se aplica global a lo que queda
                    subtotal_item -= subtotal_item * porc_global_efectivo
                else:
                    # Item tiene desc propio, NO acumulable, Y hay desc global -> Alerta! (Se procesará en Confirmar, aquí mostramos flag visual)
                    str_desc += " (!)"
                    
            if subtotal_item < 0: subtotal_item = 0.0
            
            total_monto += subtotal_item
            
            tags = []
            if p["stock"] < cant:
                tags.append("stock_insuficiente")
                
            self.tree.insert("", "end", iid=str(pid),
                             values=(pid, p["nombre"], fmt_qty(p["stock"], p.get("por_peso")), f"${precio:,.2f}", fmt_qty(cant, p.get("por_peso")), str_desc, f"${subtotal_bruto:,.2f}", f"${subtotal_item:,.2f}", nota),
                             tags=tuple(tags))
                             
        self.tree.tag_configure("stock_insuficiente", foreground=DANGER)
        self._total_final_vuelto = total_monto
        self.lbl_total.config(text=f"Monto Final: ${total_monto:,.2f}")
        self._calc_vuelto()

    def _calc_vuelto(self):
        try:
            paga = float(self.v_paga_con.get().replace(",", "."))
            if paga >= self._total_final_vuelto:
                vuelto = paga - self._total_final_vuelto
                self.lbl_vuelto.config(text=f"Vuelto: ${vuelto:,.2f}")
            else:
                self.lbl_vuelto.config(text="Vuelto: -")
        except:
            self.lbl_vuelto.config(text="Vuelto: -")

    def _on_key_enter(self, event):
        sel = self.tree.selection()
        if not sel: return "break"
        iid = sel[0]
        self._edit_item(iid)
        return "break"

    def _on_double_click(self, event):
        region = self.tree.identify_region(event.x, event.y)
        if region != "cell": return
        iid = self.tree.identify_row(event.y)
        if not iid: return
        self._edit_item(iid)
        
    def _edit_item(self, iid):
        pid = int(iid)
        data = self.items[pid]
        
        dlg = ModificarItemDialog(self, data["prod"], 
                                  cant_inicial=data["cant"], nota_inicial=data["nota"],
                                  desc_val=data.get("desc_val", 0), desc_tipo=data.get("desc_tipo", "%"), desc_acum=data.get("desc_acum", True))
                                  
        if getattr(dlg, "resultado", False) and isinstance(dlg.resultado, dict):
            self.items[pid]["cant"] = dlg.resultado["cant"]
            self.items[pid]["nota"] = dlg.resultado["nota"]
            self.items[pid]["desc_val"] = dlg.resultado["desc_val"]
            self.items[pid]["desc_tipo"] = dlg.resultado["desc_tipo"]
            self.items[pid]["desc_acum"] = dlg.resultado["desc_acum"]
            self._refresh_list()
            
    def _quitar_item(self):
        sel = self.tree.selection()
        if not sel: return
        iid = sel[0]
        pid = int(iid)
        if pid in self.items:
            self.items.pop(pid)
            self._refresh_list()

    def _registrar_todos(self):
        if not self.items:
            messagebox.showwarning("Vacío", "No hay elementos en la lista para dar salida.", parent=self)
            return

        # Calcular descuentos
        try:
            glob_val = float(self.v_desc_global_val.get().replace(",", "."))
        except:
            glob_val = 0.0
        glob_tipo = self.v_desc_global_tipo.get()
        
        total_previo = sum(float(d["prod"].get("precio", 0) or 0) * float(d["cant"]) for d in self.items.values())
        
        porc_global_efectivo = 0.0
        if glob_val > 0:
            if glob_tipo == "%":
                porc_global_efectivo = glob_val / 100.0
            elif total_previo > 0:
                porc_global_efectivo = glob_val / total_previo

        conflicto = False
        if porc_global_efectivo > 0:
            for d in self.items.values():
                if float(d.get("desc_val", 0)) > 0 and not d.get("desc_acum", True):
                    conflicto = True
                    break

        resolver_conflicto = "individual"
        if conflicto:
            resp = messagebox.askyesno(
                "Conflicto de Descuentos",
                "Existen productos con descuentos individuales NO acumulables y hay un descuento Global en el carrito.\n\n"
                "¿Desea que el descuento GLOBAL reemplace a los individuales en esos productos?\n"
                "(Sí = Usar Global, No = Priorizar Individual)",
                parent=self
            )
            resolver_conflicto = "global" if resp else "individual"

        # Generar un ID de grupo simple incremental
        grupo_id = db.get_incremental_order()

        # FASE 1: Validación y cálculo
        movimientos_a_registrar = []
        
        for pid, data in self.items.items():
            p = data["prod"]
            cant = float(data["cant"]) if p.get("por_peso", 0) else int(float(str(data["cant"])))
            nota = str(data["nota"])
            stock_actual = float(p.get("stock", 0)) if p.get("por_peso", 0) else int(float(p.get("stock", 0)))
            nombre = str(p.get("nombre", ""))
            
            # Recalcular precio para DB
            precio = float(p.get("precio", 0) or 0)
            subtotal_bruto = precio * cant
            
            desc_ind_val = float(data.get("desc_val", 0))
            desc_ind_tipo = data.get("desc_tipo", "%")
            acumulable = data.get("desc_acum", True)
            
            subtotal_item = subtotal_bruto
            aplicar_ind = False
            aplicar_glob = False
            
            if desc_ind_val > 0:
                if porc_global_efectivo > 0 and not acumulable:
                    if resolver_conflicto == "global":
                        aplicar_glob = True
                    else:
                        aplicar_ind = True
                else:
                    aplicar_ind = True
                    if porc_global_efectivo > 0:
                        aplicar_glob = True
            elif porc_global_efectivo > 0:
                aplicar_glob = True
                
            if aplicar_ind:
                if desc_ind_tipo == "%":
                    subtotal_item -= subtotal_bruto * (desc_ind_val / 100.0)
                else:
                    subtotal_item -= desc_ind_val
                    
            if aplicar_glob:
                subtotal_item -= subtotal_item * porc_global_efectivo
                
            if subtotal_item < 0: subtotal_item = 0.0
            
            # Replicar la logica de control de stock
            forzar_salida = False
            if stock_actual < cant:
                resp = messagebox.askyesno(
                    "⚠️ Stock insuficiente",
                    f"Para el producto '{nombre}':\n"
                    f"Stock actual: {stock_actual}  |  Cantidad solicitada: {cant}\n\n"
                    f"¿Desea registrar la salida de todas formas para este artículo?\n"
                    f"(Se marcará como excepción en el historial)",
                    icon="warning", parent=self
                )
                if resp:
                    forzar_salida = True
                    nota = (nota + " [EXCEPCIÓN: stock insuficiente]").strip()
                else:
                    return # Abort confirmation and keep window open for corrections
                    
            movimientos_a_registrar.append({
                "pid": pid,
                "cant": cant,
                "nota": nota,
                "forzar": forzar_salida,
                "precio_total": subtotal_item
            })

        # FASE 2: Registro en Base de Datos
        for mov in movimientos_a_registrar:
            db.registrar_movimiento(
                mov["pid"], 
                "salida", 
                mov["cant"], 
                mov["nota"], 
                forzar=mov["forzar"], 
                precio_total=mov["precio_total"], 
                grupo_id=grupo_id
            )

        self.resultado = True
        self.destroy()

# ──────────────────────────────────────────────
#  Ventana de Resumen de Importación
# ──────────────────────────────────────────────

class ImportacionResumenDialog(tk.Toplevel):
    def __init__(self, parent, nombres_nuevos, nombres_actualizados):
        super().__init__(parent)
        self.title("Resumen de Importación")
        self.configure(bg=BG2)
        self.bind("<Escape>", lambda e: self.destroy())
        self.geometry("600x450")
        self.resizable(False, False)
        self.grab_set()

        frame = tk.Frame(self, bg=BG2, padx=20, pady=20)
        frame.pack(fill="both", expand=True)

        tk.Label(frame, text="✅ Importación Completada", bg=BG2, fg=SUCCESS, font=("Segoe UI", 14, "bold")).pack(pady=(0, 10))
        
        lbl_nuevos = tk.Label(frame, text=f"➕ Nuevos productos agregados: {len(nombres_nuevos)}", bg=BG2, fg=TEXT, font=("Segoe UI", 11, "bold"))
        lbl_nuevos.pack(anchor="w", pady=5)
        
        if nombres_nuevos:
            list_nuevos_frame = tk.Frame(frame, bg=BG, bd=1, relief="solid")
            list_nuevos_frame.pack(fill="both", expand=True, pady=(0, 10))
            
            scrollbar_n = ttk.Scrollbar(list_nuevos_frame)
            scrollbar_n.pack(side="right", fill="y")
            
            listbox_n = tk.Listbox(list_nuevos_frame, yscrollcommand=scrollbar_n.set, bg=BG, fg=TEXT, height=4,
                                 font=("Segoe UI", 9), relief="flat", highlightthickness=0, selectbackground=ACCENT)
            listbox_n.pack(side="left", fill="both", expand=True, padx=2, pady=2)
            scrollbar_n.config(command=listbox_n.yview)
            
            for nombre in nombres_nuevos:
                listbox_n.insert("end", f"• {nombre}")

        lbl_act = tk.Label(frame, text=f"🔄 Productos actualizados: {len(nombres_actualizados)}", bg=BG2, fg=TEXT, font=("Segoe UI", 11, "bold"))
        lbl_act.pack(anchor="w", pady=5)

        if nombres_actualizados:
            list_frame = tk.Frame(frame, bg=BG, bd=1, relief="solid")
            list_frame.pack(fill="both", expand=True, pady=(0, 5))
            
            scrollbar = ttk.Scrollbar(list_frame)
            scrollbar.pack(side="right", fill="y")
            
            listbox = tk.Listbox(list_frame, yscrollcommand=scrollbar.set, bg=BG, fg=TEXT, 
                                 font=("Segoe UI", 9), relief="flat", highlightthickness=0, selectbackground=ACCENT)
            listbox.pack(side="left", fill="both", expand=True, padx=2, pady=2)
            scrollbar.config(command=listbox.yview)
            
            for nombre, cambios in nombres_actualizados:
                cambios_str = " | ".join(cambios)
                listbox.insert("end", f"• {nombre} ({cambios_str})")

        btn_frame = tk.Frame(self, bg=BG2, pady=15)
        btn_frame.pack(side="bottom", fill="x")
        styled_btn(btn_frame, "Aceptar", self.destroy, color=BG3).pack()


class ILovePdfDialog(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Convertidor PDF a Excel")
        self.geometry("450x250")
        self.configure(bg=BG)
        self.transient(parent)

        lbl_title = tk.Label(self, text="Convertir lista de precios PDF a Excel", bg=BG, fg=TEXT, font=("Segoe UI", 12, "bold"))
        lbl_title.pack(pady=15)

        tk.Label(self, text="Selecciona un PDF de tu proveedor para extraer\nlos productos y precios a un archivo Excel.", bg=BG, fg=TEXT_DIM, wraplength=400).pack(pady=5)

        btn_seleccionar = styled_btn(self, "📂 Seleccionar PDF y Convertir", self.convertir_pdf, color=ACCENT)
        btn_seleccionar.pack(pady=15)

        self.lbl_status = tk.Label(self, text="", bg=BG, fg=TEXT_DIM)
        self.lbl_status.pack(pady=5)
        
    def convertir_pdf(self):
        filepath = filedialog.askopenfilename(
            filetypes=[("Archivos PDF", "*.pdf")],
            title="Seleccionar lista en PDF",
            parent=self
        )
        if not filepath:
            return
        import pathlib
        savepath = str(pathlib.Path(filepath).with_name(pathlib.Path(filepath).stem + " - Convertido.xlsx"))
        
        self.lbl_status.config(text="Subiendo a iLovePDF y procesando... Esto puede tardar unos segundos.", fg=WARNING)
        self.update()
        
        def run_conversion():
            try:
                import pdfplumber
                import pandas as pd
                import os

                all_data = []
                with pdfplumber.open(filepath) as pdf:
                    for page in pdf.pages:
                        tables = page.extract_tables()
                        for table in tables:
                            for row in table:
                                clean_row = [str(cell).strip() if cell is not None else '' for cell in row]
                                if not any(clean_row):
                                    continue
                                first = clean_row[0].upper()
                                if any(skip in first for skip in ['PRODUCTOS', 'LISTA DE PRECIOS', 'DISTRIBUIDORA', 'COMPRAS SUPERIORES', 'POR CUALQUIER']):
                                    continue
                                all_data.append(clean_row)

                if not all_data:
                    raise Exception('No se encontraron datos tabulares en el PDF.')

                max_cols = max(len(row) for row in all_data)
                standardized = [row + [''] * (max_cols - len(row)) for row in all_data]
                columns = [f'Columna_{i+1}' for i in range(max_cols)]
                if max_cols >= 4:
                    columns[0] = 'Producto'
                    columns[1] = 'Precio'
                    columns[3] = 'Marca_o_Cat'

                df = pd.DataFrame(standardized, columns=columns)
                df.replace('', pd.NA, inplace=True)
                df.dropna(axis=1, how='all', inplace=True)
                df.fillna('', inplace=True)
                df.to_excel(savepath, index=False)

                self.after(0, lambda: self.deiconify())
                self.after(0, lambda: self.lift())
                self.after(0, lambda: self.focus_force())
                self.after(0, lambda: self.lbl_status.config(text=f'Conversión exitosa! {len(df)} filas extraídas.', fg=SUCCESS))
                self.after(0, lambda: messagebox.showinfo('Listo',
                    f'El archivo Excel ha sido guardado en:\n{savepath}\n\n{len(df)} filas extraídas.', parent=self))
            except Exception as e:
                self.after(0, lambda: self.deiconify())
                self.after(0, lambda: self.lift())
                self.after(0, lambda: self.focus_force())
                self.after(0, lambda: self.lbl_status.config(text='Error en la conversión.', fg=DANGER))
                self.after(0, lambda e=e: messagebox.showerror('Error',
                    f'Ocurrió un error:\n{str(e)}', parent=self))

        import threading
        threading.Thread(target=run_conversion, daemon=True).start()


class GenerarListaPdfDialog(tk.Toplevel):
    """Diálogo para generar lista de precios en PDF."""
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Generar Lista de Precios PDF")
        self.geometry("420x380")
        self.configure(bg=BG)
        self.transient(parent)
        self.resizable(False, False)

        tk.Label(self, text="Generar Lista de Precios PDF", bg=BG, fg=TEXT,
                 font=("Segoe UI", 13, "bold")).pack(pady=(15, 5))
        tk.Label(self, text="Configura las opciones de tu lista:", bg=BG,
                 fg=TEXT_DIM).pack(pady=(0, 10))

        # --- Orden ---
        frame_orden = tk.LabelFrame(self, text="Ordenar por", bg=BG, fg=TEXT,
                                     font=("Segoe UI", 10, "bold"), padx=10, pady=8)
        frame_orden.pack(fill="x", padx=20, pady=5)

        self.var_orden = tk.StringVar(value="alfabetico")
        opciones_orden = [
            ("Alfabético (A-Z)", "alfabetico"),
            ("Por Categoría", "categoria"),
            ("Por Marca", "marca"),
            ("Por Precio (menor a mayor)", "precio"),
        ]
        for texto, valor in opciones_orden:
            rb = tk.Radiobutton(frame_orden, text=texto, variable=self.var_orden,
                                value=valor, bg=BG, fg=TEXT, selectcolor=BG2,
                                activebackground=BG, activeforeground=TEXT,
                                font=("Segoe UI", 9))
            rb.pack(anchor="w")

        # --- Incluir sin stock ---
        frame_stock = tk.Frame(self, bg=BG)
        frame_stock.pack(fill="x", padx=20, pady=10)

        self.var_sin_stock = tk.BooleanVar(value=False)
        cb = tk.Checkbutton(frame_stock, text="Incluir productos sin stock",
                            variable=self.var_sin_stock, bg=BG, fg=TEXT,
                            selectcolor=BG2, activebackground=BG,
                            activeforeground=TEXT, font=("Segoe UI", 10))
        cb.pack(anchor="w")

        # --- Botón generar ---
        styled_btn(self, "📄 Generar PDF", self._generar, color=ACCENT).pack(pady=15)

        self.lbl_status = tk.Label(self, text="", bg=BG, fg=TEXT_DIM)
        self.lbl_status.pack()

    def _generar(self):
        import threading
        self.lbl_status.config(text="Generando PDF...", fg=WARNING)
        self.update()
        threading.Thread(target=self._generar_pdf_thread, daemon=True).start()

    def _generar_pdf_thread(self):
        try:
            from fpdf import FPDF
            import os, sys
            from pathlib import Path
            from datetime import datetime

            # --- Obtener productos ---
            productos = db.get_productos()

            # Filtrar sin stock si corresponde
            if not self.var_sin_stock.get():
                productos = [p for p in productos if (p.get("stock") or 0) > 0]

            if not productos:
                self.after(0, lambda: messagebox.showwarning("Sin datos",
                    "No hay productos para incluir en la lista.", parent=self))
                self.after(0, lambda: self.lbl_status.config(text="", fg=TEXT_DIM))
                return

            # --- Ordenar ---
            orden = self.var_orden.get()
            if orden == "alfabetico":
                productos.sort(key=lambda p: (p.get("nombre") or "").lower())
            elif orden == "categoria":
                productos.sort(key=lambda p: ((p.get("categoria") or "Sin Categoría").lower(),
                                              (p.get("nombre") or "").lower()))
            elif orden == "marca":
                productos.sort(key=lambda p: ((p.get("marca") or "Sin Marca").lower(),
                                              (p.get("nombre") or "").lower()))
            elif orden == "precio":
                productos.sort(key=lambda p: p.get("precio") or 0)

            # --- Logo path ---
            if getattr(sys, 'frozen', False):
                base_dir = Path(sys._MEIPASS)
            else:
                base_dir = Path(__file__).parent
            logo_path = str(base_dir / "hornero_logo.png")
            if not os.path.exists(logo_path):
                logo_path = None

            # --- Colores del estilo Profesional Clásico ---
            HEADER_R, HEADER_G, HEADER_B = 44, 62, 80       # Azul oscuro
            ACCENT_R, ACCENT_G, ACCENT_B = 52, 152, 219     # Azul acento
            ALT_ROW_R, ALT_ROW_G, ALT_ROW_B = 235, 240, 245 # Gris claro alterno
            CAT_BG_R, CAT_BG_G, CAT_BG_B = 52, 73, 94       # Fondo categoría

            # --- Crear PDF ---
            pdf = FPDF()
            pdf.set_auto_page_break(auto=True, margin=20)
            pdf.add_page()
            page_w = pdf.w - 20  # margen izq 10 + der 10

            # --- Encabezado ---
            # Franja superior
            pdf.set_fill_color(HEADER_R, HEADER_G, HEADER_B)
            pdf.rect(0, 0, 210, 38, 'F')

            # Logo
            if logo_path:
                pdf.image(logo_path, x=12, y=5, w=28)

            # Título
            pdf.set_xy(44, 8)
            pdf.set_text_color(255, 255, 255)
            pdf.set_font("Helvetica", "B", 18)
            pdf.cell(0, 10, "Lista Jaulas Gonzalez", ln=True)

            # Fecha
            pdf.set_xy(44, 20)
            pdf.set_font("Helvetica", "", 10)
            fecha = datetime.now().strftime("%d/%m/%Y")
            pdf.cell(0, 8, f"Fecha: {fecha}", ln=True)

            pdf.ln(18)

            # --- Encabezado de tabla ---
            col_widths = [page_w * 0.42, page_w * 0.22, page_w * 0.20, page_w * 0.16]
            headers = ["Producto", "Categoría", "Marca", "Precio"]

            def draw_table_header():
                pdf.set_fill_color(ACCENT_R, ACCENT_G, ACCENT_B)
                pdf.set_text_color(255, 255, 255)
                pdf.set_font("Helvetica", "B", 10)
                pdf.set_draw_color(255, 255, 255)
                for i, h in enumerate(headers):
                    align = "R" if i == 3 else "L"
                    pdf.cell(col_widths[i], 9, f"  {h}" if i < 3 else f"{h}  ",
                             border="LR", fill=True, align=align)
                pdf.ln()

            draw_table_header()

            # --- Filas de productos ---
            pdf.set_font("Helvetica", "", 8)

            for idx, p in enumerate(productos):
                # Check page space
                if pdf.get_y() > 272:
                    pdf.add_page()
                    draw_table_header()

                # Fila alternada
                is_alt = idx % 2 == 0
                if is_alt:
                    pdf.set_fill_color(ALT_ROW_R, ALT_ROW_G, ALT_ROW_B)
                else:
                    pdf.set_fill_color(255, 255, 255)

                pdf.set_text_color(33, 33, 33)

                nombre = (p.get("nombre") or "")[:65]
                categoria = (p.get("categoria") or "")[:25]
                marca = (p.get("marca") or "")[:25]
                precio_val = p.get("precio") or 0
                precio_str = f"${precio_val:,.2f}"

                pdf.set_draw_color(220, 220, 220)
                pdf.cell(col_widths[0], 7, f"  {nombre}", border="LR", fill=True, align="L")
                pdf.cell(col_widths[1], 7, f"  {categoria}", border="LR", fill=True, align="L")
                pdf.cell(col_widths[2], 7, f"  {marca}", border="LR", fill=True, align="L")
                pdf.set_font("Helvetica", "B", 8)
                pdf.cell(col_widths[3], 7, f"{precio_str}  ", border="LR", fill=True, align="R")
                pdf.set_font("Helvetica", "", 8)
                pdf.ln()

            # --- Pie de página ---
            pdf.ln(5)
            pdf.set_draw_color(ACCENT_R, ACCENT_G, ACCENT_B)
            pdf.line(10, pdf.get_y(), 200, pdf.get_y())
            pdf.ln(3)
            pdf.set_text_color(120, 120, 120)
            pdf.set_font("Helvetica", "I", 8)
            pdf.cell(0, 5, f"Lista generada el {fecha}  |  {len(productos)} productos",
                     align="C")

            # --- Guardar ---
            savepath = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("Archivo PDF", "*.pdf")],
                initialfile=f"Lista Jaulas Gonzalez - {fecha.replace('/', '-')}.pdf",
                title="Guardar lista de precios",
                parent=self
            )
            if not savepath:
                self.after(0, lambda: self.lbl_status.config(text="Cancelado.", fg=TEXT_DIM))
                return

            pdf.output(savepath)

            self.after(0, lambda: self.deiconify())
            self.after(0, lambda: self.lift())
            self.after(0, lambda: self.focus_force())
            self.after(0, lambda: self.lbl_status.config(
                text=f"PDF generado con {len(productos)} productos.", fg=SUCCESS))
            self.after(0, lambda: messagebox.showinfo("Listo",
                f"Lista de precios guardada en:\n{savepath}\n\n{len(productos)} productos incluidos.",
                parent=self))
        except Exception as e:
            self.after(0, lambda: self.deiconify())
            self.after(0, lambda: self.lift())
            self.after(0, lambda: self.focus_force())
            self.after(0, lambda: self.lbl_status.config(text="Error al generar.", fg=DANGER))
            self.after(0, lambda e=e: messagebox.showerror("Error",
                f"Error al generar el PDF:\n{str(e)}", parent=self))

class ReporteCapitalDialog(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Reporte de Capital Total")
        self.geometry("900x600")
        self.minsize(700, 500)
        self.configure(bg=BG)
        self.grab_set()
        
        self.productos = db.get_productos()
        
        self.categorias = sorted(list({p.get("categoria") or "Sin Categoría" for p in self.productos}))
        self.marcas = sorted(list({p.get("marca") or "Sin Marca" for p in self.productos}))
        
        self.excluidos_ids = set()
        
        self._build_ui()
        self._calcular()
        
    def _build_ui(self):
        # Header
        header = tk.Frame(self, bg=BG2, pady=20)
        header.pack(fill="x")
        
        # Split header into two columns
        h_left = tk.Frame(header, bg=BG2)
        h_left.pack(side="left", expand=True)
        tk.Label(h_left, text="Capital Invertido (Costo)", bg=BG2, fg=TEXT_DIM, font=("Segoe UI", 12)).pack()
        self.lbl_costo = tk.Label(h_left, text="$ 0.00", bg=BG2, fg=SUCCESS, font=("Segoe UI", 24, "bold"))
        self.lbl_costo.pack()

        h_right = tk.Frame(header, bg=BG2)
        h_right.pack(side="left", expand=True)
        tk.Label(h_right, text="Ganancia Potencial (Venta)", bg=BG2, fg=TEXT_DIM, font=("Segoe UI", 12)).pack()
        self.lbl_venta = tk.Label(h_right, text="$ 0.00", bg=BG2, fg=WARNING, font=("Segoe UI", 24, "bold"))
        self.lbl_venta.pack()
        
        # Body
        body = tk.Frame(self, bg=BG)
        body.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Col 1: Categorías
        col_cat = tk.Frame(body, bg=BG)
        col_cat.pack(side="left", fill="both", expand=True, padx=10)
        tk.Label(col_cat, text="Categorías a Incluir", bg=BG, fg=TEXT, font=("Segoe UI", 10, "bold")).pack(anchor="w")
        tk.Button(col_cat, text="Todas / Ninguna", command=self._toggle_cat, bg=BG3, fg=TEXT, relief="flat", font=("Segoe UI", 8), cursor="hand2").pack(anchor="w", pady=2)
        
        cat_frame = tk.Frame(col_cat, bg=BG2)
        cat_frame.pack(fill="both", expand=True)
        self.lb_cat = tk.Listbox(cat_frame, selectmode="multiple", bg=BG2, fg=TEXT, font=("Segoe UI", 10), relief="flat", selectbackground=ACCENT, exportselection=False)
        self.lb_cat.pack(side="left", fill="both", expand=True)
        for c in self.categorias:
            self.lb_cat.insert("end", c)
            self.lb_cat.selection_set("end")
        self.lb_cat.bind("<<ListboxSelect>>", lambda e: self.after(10, self._calcular))
        
        # Col 2: Marcas
        col_marca = tk.Frame(body, bg=BG)
        col_marca.pack(side="left", fill="both", expand=True, padx=10)
        tk.Label(col_marca, text="Marcas a Incluir", bg=BG, fg=TEXT, font=("Segoe UI", 10, "bold")).pack(anchor="w")
        tk.Button(col_marca, text="Todas / Ninguna", command=self._toggle_marca, bg=BG3, fg=TEXT, relief="flat", font=("Segoe UI", 8), cursor="hand2").pack(anchor="w", pady=2)
        
        marca_frame = tk.Frame(col_marca, bg=BG2)
        marca_frame.pack(fill="both", expand=True)
        self.lb_marca = tk.Listbox(marca_frame, selectmode="multiple", bg=BG2, fg=TEXT, font=("Segoe UI", 10), relief="flat", selectbackground=ACCENT, exportselection=False)
        self.lb_marca.pack(side="left", fill="both", expand=True)
        for m in self.marcas:
            self.lb_marca.insert("end", m)
            self.lb_marca.selection_set("end")
        self.lb_marca.bind("<<ListboxSelect>>", lambda e: self.after(10, self._calcular))

        # Col 3: Excluidos
        col_prod = tk.Frame(body, bg=BG)
        col_prod.pack(side="left", fill="both", expand=True, padx=10)
        
        tk.Label(col_prod, text="Productos Individuales a Excluir", bg=BG, fg=TEXT, font=("Segoe UI", 10, "bold")).pack(anchor="w")
        tk.Label(col_prod, text="(Escribe para buscar. Doble clic para excluir)", bg=BG, fg=TEXT_DIM, font=("Segoe UI", 8)).pack(anchor="w")
        
        search_f = tk.Frame(col_prod, bg=BG)
        search_f.pack(fill="x", pady=5)
        self.v_buscar = tk.StringVar()
        entry_search = tk.Entry(search_f, textvariable=self.v_buscar, font=("Segoe UI", 10), bg=BG2, fg=TEXT, insertbackground=TEXT)
        entry_search.pack(side="left", fill="x", expand=True)
        self.v_buscar.trace_add("write", self._on_search)
        
        self.lb_search = tk.Listbox(col_prod, height=6, bg=BG2, fg=TEXT, font=("Segoe UI", 9), relief="flat", selectbackground=WARNING, exportselection=False)
        self.lb_search.pack(fill="x")
        self.lb_search.bind("<Double-Button-1>", self._add_excluded_prod)
        
        tk.Label(col_prod, text="Lista de Excluidos (Doble clic para quitar):", bg=BG, fg=TEXT_DIM, font=("Segoe UI", 9)).pack(anchor="w", pady=(10,0))
        self.lb_excluidos = tk.Listbox(col_prod, bg=BG2, fg=DANGER, font=("Segoe UI", 10), relief="flat", selectbackground=BG3, exportselection=False)
        self.lb_excluidos.pack(fill="both", expand=True)
        self.lb_excluidos.bind("<Double-Button-1>", self._remove_excluded_prod)
        
        # Bottom
        foot = tk.Frame(self, bg=BG, pady=10)
        foot.pack(side="bottom", fill="x")
        styled_btn(foot, "Cerrar Panel", self.destroy, color=BG3).pack()

        self._on_search()

    def _toggle_cat(self):
        if len(self.lb_cat.curselection()) > 0:
            self.lb_cat.selection_clear(0, "end")
        else:
            self.lb_cat.selection_set(0, "end")
        self.after(10, self._calcular)

    def _toggle_marca(self):
        if len(self.lb_marca.curselection()) > 0:
            self.lb_marca.selection_clear(0, "end")
        else:
            self.lb_marca.selection_set(0, "end")
        self.after(10, self._calcular)
        
    def _on_search(self, *args):
        q = self.v_buscar.get().strip().lower()
        self.lb_search.delete(0, "end")
        self._search_map = []
        if len(q) < 2:
            return
        
        count = 0
        for p in self.productos:
            if p["id"] in self.excluidos_ids: continue
            if q in p["nombre"].lower() or q in str(p.get("codigo", "")).lower():
                self.lb_search.insert("end", f"{p['nombre']} (Stock: {p['stock']})")
                self._search_map.append(p)
                count += 1
                if count >= 10: break

    def _add_excluded_prod(self, event):
        sel = self.lb_search.curselection()
        if not sel: return
        p = self._search_map[sel[0]]
        self.excluidos_ids.add(p["id"])
        self.lb_excluidos.insert("end", p["nombre"])
        self._on_search()
        self._calcular()

    def _remove_excluded_prod(self, event):
        sel = self.lb_excluidos.curselection()
        if not sel: return
        idx = sel[0]
        nombre = self.lb_excluidos.get(idx)
        self.lb_excluidos.delete(idx)
        for p in self.productos:
            if p["nombre"] == nombre and p["id"] in self.excluidos_ids:
                self.excluidos_ids.remove(p["id"])
                break
        self._on_search()
        self._calcular()

    def _calcular(self):
        sel_cat_idx = self.lb_cat.curselection()
        sel_cats = {self.categorias[i] for i in sel_cat_idx}
        
        sel_marca_idx = self.lb_marca.curselection()
        sel_marcas = {self.marcas[i] for i in sel_marca_idx}
        
        total_venta = 0.0
        total_costo = 0.0
        for p in self.productos:
            if p["id"] in self.excluidos_ids: continue
            
            c = p.get("categoria") or "Sin Categoría"
            m = p.get("marca") or "Sin Marca"
            
            if c not in sel_cats: continue
            if m not in sel_marcas: continue
            
            stock = float(p["stock"] or 0)
            precio = float(p["precio"] or 0)
            precio_costo = float(p.get("precio_costo", 0) or 0)
            
            if stock > 0:
                if precio > 0:
                    total_venta += stock * precio
                if precio_costo > 0:
                    total_costo += stock * precio_costo
                
        self.lbl_costo.config(text=f"$ {total_costo:,.2f}")
        self.lbl_venta.config(text=f"$ {total_venta:,.2f}")

# ──────────────────────────────────────────────
#  Ventana de Importación (Asistente de Mapeo)
# ──────────────────────────────────────────────

class ImportarDialog(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Asistente de Importación Avanzado")
        self.configure(bg=BG2)
        self.bind("<Escape>", lambda e: self.destroy())
        self.geometry("680x480")
        self.grab_set()
        
        self.filepath = None
        self.mapeo_resultado = {}  # {campo: {'col': 'A', 'inicio': 7, 'fin': None}}
        self.vars = {}  # {campo: {'col': tk.StringVar(), 'inicio': tk.StringVar(), 'fin': tk.StringVar()}}
        
        self._build_ui()
        self.wait_window()

    def _build_ui(self):
        # Paso 1: Seleccionar archivo
        frame_top = tk.Frame(self, bg=BG2, pady=15, padx=20)
        frame_top.pack(fill="x")
        
        tk.Label(frame_top, text="1. Seleccionar archivo", bg=BG2, fg=ACCENT2, font=("Segoe UI", 11, "bold")).pack(anchor="w")
        
        file_frame = tk.Frame(frame_top, bg=BG2)
        file_frame.pack(fill="x", pady=5)
        
        self.lbl_file = tk.Label(file_frame, text="Ningún archivo seleccionado", bg=BG2, fg=TEXT_DIM, font=("Segoe UI", 9))
        self.lbl_file.pack(side="left")
        
        styled_btn(file_frame, "📂 Examinar", self._seleccionar_archivo, color=BG3).pack(side="right")
        
        # Paso 2: Configuración del Layout de Coordenadas
        self.frame_mapeo = tk.Frame(self, bg=BG2, pady=10, padx=20)
        
        tk.Label(self.frame_mapeo, text="2. Mapear datos por coordenadas", bg=BG2, fg=ACCENT2, font=("Segoe UI", 11, "bold")).pack(anchor="w", pady=(0, 10))
        tk.Label(self.frame_mapeo, text="Deja vacía la columna de los campos que no deseas importar.", bg=BG2, fg=TEXT_DIM, font=("Segoe UI", 8)).pack(anchor="w", pady=(0, 5))
        
        self.mapeo_container = tk.Frame(self.frame_mapeo, bg=BG3, padx=15, pady=15)
        self.mapeo_container.pack(fill="both", expand=True)
        
        # Headers para la tabla
        tk.Label(self.mapeo_container, text="Campo del Sistema", bg=BG3, fg=TEXT_DIM).grid(row=0, column=0, sticky="w", padx=5)
        tk.Label(self.mapeo_container, text="Columna (Letra)", bg=BG3, fg=TEXT_DIM).grid(row=0, column=1, padx=5)
        tk.Label(self.mapeo_container, text="Fila Inicio", bg=BG3, fg=TEXT_DIM).grid(row=0, column=2, padx=5)
        tk.Label(self.mapeo_container, text="Fila Fin (Opcional)", bg=BG3, fg=TEXT_DIM).grid(row=0, column=3, padx=5)
        
        # Botones inferiores
        self.frame_btns = tk.Frame(self, bg=BG2, pady=15)
        self.frame_btns.pack(side="bottom", fill="x")
        
        self.btn_importar = styled_btn(self.frame_btns, "✔ Iniciar Importación", self._procesar, color=SUCCESS)

    def _seleccionar_archivo(self):
        ruta = filedialog.askopenfilename(
            filetypes=[("Excel", "*.xlsx *.xls"), ("CSV", "*.csv")],
            title="Seleccionar archivo de productos"
        )
        if not ruta:
            return
            
        self.filepath = ruta
        self.lbl_file.config(text=self.filepath.split("/")[-1], fg=TEXT)
        self._mostrar_mapeo()

    def _mostrar_mapeo(self):
        # Limpiar contenedor previo (menos los headers de la fila 0)
        for w in self.mapeo_container.winfo_children():
            if int(w.grid_info().get("row", 0)) > 0:
                w.destroy()
            
        self.frame_mapeo.pack(fill="both", expand=True)
        self.btn_importar.pack(side="right", padx=20)
        
        campos_sistema = [
            ("nombre",       "Nombre (Requerido)", True, "", "", ""),
            ("codigo",       "Código", False, "", "", ""),
            ("categoria",    "Categoría", False, "", "", ""),
            ("marca",        "Marca", False, "", "", ""),
            ("precio_costo", "Precio Costo", False, "", "", ""),
            ("precio",       "Precio Venta", False, "", "", ""),
            ("stock",        "Stock Inicial", False, "", "", ""),
            ("minimo",       "Stock Mínimo", False, "", "", "")
        ]
        
        for i, (key, label, req, d_col, d_in, d_fin) in enumerate(campos_sistema, start=1):
            # Etiqueta
            tk.Label(self.mapeo_container, text=label, bg=BG3, fg=TEXT if req else TEXT_DIM, 
                     font=("Segoe UI", 9, "bold" if req else "normal")).grid(row=i, column=0, sticky="w", pady=6, padx=5)
            
            # Variables y tracking
            v_col = tk.StringVar(value=d_col)
            v_ini = tk.StringVar(value=d_in)
            v_fin = tk.StringVar(value=d_fin)
            self.vars[key] = {"col": v_col, "inicio": v_ini, "fin": v_fin}
            
            # Entrys
            tk.Entry(self.mapeo_container, textvariable=v_col, width=8, bg=BG2, fg=TEXT, justify="center").grid(row=i, column=1, padx=10)
            tk.Entry(self.mapeo_container, textvariable=v_ini, width=8, bg=BG2, fg=TEXT, justify="center").grid(row=i, column=2, padx=10)
            tk.Entry(self.mapeo_container, textvariable=v_fin, width=12, bg=BG2, fg=TEXT, justify="center").grid(row=i, column=3, padx=10)

    def _procesar(self):
        mapeo = {}
        for key, config in self.vars.items():
            c = config['col'].get().strip().upper()
            if c:  # Solo mapeamos si el usuario ingresó una columna
                mapeo[key] = {
                    "col": c,
                    "inicio": config['inicio'].get().strip() or "1",
                    "fin": config['fin'].get().strip() or None
                }
                
        if "nombre" not in mapeo:
            messagebox.showwarning("Faltan campos", "Debes mapear obligatoriamente la columna para Nombre.", parent=self)
            return
            
        self.mapeo_resultado = mapeo
        self.destroy()



# ──────────────────────────────────────────────
#  Aplicación principal
# ──────────────────────────────────────────────

class StockApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Hornerito Forrajería")
        self.geometry("1100x680")
        self.minsize(900, 560)
        try:
            self.state("zoomed")
        except tk.TclError:
            pass
        self.configure(bg=BG)
        self._setup_styles()
        self._seleccionados = set()
        self._seleccionados_pos = set()
        self._last_selected_iid = None
        self._drag_start_iid = None
        self._drag_start_action = None
        self._cantidades_carrito = {}
        self.v_marca_filtro = None
        self.cmb_marca = None
        
        # Estado del modo Stock Rápido
        self._modo_stock_rapido = False
        self._cambios_stock_tmp = {}
        self._undo_stock_stack = []
        self._stock_entry_active = None

        self._build_ui()
        self.refresh_productos()
        
        # Foco inicial en la barra de búsqueda para que funcionen los atajos y escritura directa
        self.entry_buscar.focus_set()
        
        self.protocol("WM_DELETE_WINDOW", self._on_closing)

    def _on_closing(self):
        if hasattr(self, "_cantidades_carrito") and getattr(self, "_cantidades_carrito", None) and len(self._cantidades_carrito) > 0:
            count = len(self._cantidades_carrito)
            if messagebox.askyesno("Cambios no guardados", f"Tienes {count} producto(s) en el carrito pendientes.\n\n¿Deseas cerrar la aplicación de todas formas y perder el carrito?", parent=self):
                self.destroy()
        else:
            if messagebox.askyesno("Salir", "¿Estás seguro de que deseas cerrar la aplicación?", parent=self):
                self.destroy()

    # ── Estilos ttk ─────────────────────────────

    def _setup_styles(self):
        style = ttk.Style(self)
        style.theme_use("clam")

        # Notebook (pestañas)
        style.configure("TNotebook", background=BG, borderwidth=0)
        style.configure("TNotebook.Tab",
                        background=BG2, foreground=TEXT_DIM,
                        font=("Segoe UI", 10), padding=(14, 7))
        style.map("TNotebook.Tab",
                  background=[("selected", BG3)],
                  foreground=[("selected", ACCENT2)])

        # Treeview
        style.configure("Treeview",
                        background=BG2, foreground=TEXT,
                        fieldbackground=BG2,
                        rowheight=28, font=("Segoe UI", 9))
        style.configure("Treeview.Heading",
                        background=BG3, foreground=ACCENT2,
                        font=("Segoe UI", 9, "bold"), relief="flat")
        style.map("Treeview",
                  background=[("selected", ACCENT)],
                  foreground=[("selected", TEXT)])

        # Scrollbar
        style.configure("Vertical.TScrollbar",
                        background=BG3, troughcolor=BG,
                        arrowcolor=TEXT_DIM, borderwidth=0)

    # ── Cabecera ─────────────────────────────────

    def _build_header(self, parent):
        header = tk.Frame(parent, bg=BG, pady=12)
        header.pack(fill="x", padx=16)

        tk.Label(header, text="\U0001f4e6  Control de Stock",
                 bg=BG, fg=ACCENT2,
                 font=("Segoe UI", 18, "bold")).pack(side="left")

        # ── Toggle de bloqueo de edición ──────────────
        self.v_edicion = tk.BooleanVar(value=False)  # False = bloqueado
        self.btn_lock = tk.Button(
            header,
            text="\U0001f512  Edición bloqueada (Alt+E)",
            command=self._toggle_edicion,
            bg="#8b3a3a", fg=TEXT,
            relief="flat", padx=10, pady=4,
            font=("Segoe UI", 9, "bold"),
            cursor="hand2"
        )
        self.btn_lock.pack(side="right", padx=12)

        # Ruedita de configuración
        tk.Button(
            header, text="⚙️",
            command=self._abrir_configuracion,
            bg=BG3, fg=TEXT, relief="flat",
            font=("Segoe UI", 12), padx=6, pady=4,
            cursor="hand2"
        ).pack(side="right", padx=4)
        # (Datos movidos a la pestaña de Edición)
        self.lbl_total = tk.Label(header, text="", bg=BG, fg=TEXT_DIM,
                                  font=("Segoe UI", 9))
        self.lbl_total.pack(side="right", padx=8)

    # ── UI principal ─────────────────────────────

    def _build_ui(self):
        self._build_header(self)

        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True, padx=12, pady=(0, 12))

        # Pestaña 1: Productos (POS)
        tab_productos = tk.Frame(self.notebook, bg=BG)
        self.notebook.add(tab_productos, text="  🗃  Productos (F7)  ")
        self._build_tab_productos(tab_productos)

        # Pestaña 2: Historial
        tab_historial = tk.Frame(self.notebook, bg=BG)
        self.notebook.add(tab_historial, text="  📋  Historial (F8)  ")
        self._build_tab_historial(tab_historial)

        # Pestaña 3: Alertas
        tab_alertas = tk.Frame(self.notebook, bg=BG)
        self.notebook.add(tab_alertas, text="  ⚠  Alertas (F9)  ")
        self._build_tab_alertas(tab_alertas)

        # Pestaña 4: Edición (solo accesible con edición habilitada) — va al final
        self.tab_edicion_frame = tk.Frame(self.notebook, bg=BG)
        self.notebook.add(self.tab_edicion_frame, text="  ✏️  Edición (F10)  ")
        self._build_tab_edicion(self.tab_edicion_frame)

        self.notebook.bind("<<NotebookTabChanged>>", self._on_tab_change)

        # Tab navigation shortcuts
        self.bind_all("<F7>", lambda e: self.notebook.select(0))
        self.bind_all("<F8>", lambda e: self.notebook.select(1))
        self.bind_all("<F9>", lambda e: self.notebook.select(2))
        self.bind_all("<F10>", lambda e: self.notebook.select(3))
        self.bind_all("<F5>", lambda e: self._refresh_all_tabs())
        self.bind_all("<Alt-e>", lambda e: self._toggle_edicion())
        
    def _refresh_all_tabs(self):
        self.refresh_productos()
        if hasattr(self, 'refresh_tab_edicion'):
            self.refresh_tab_edicion()
        self.refresh_historial()
        self.refresh_alertas()

    # ── Pestaña Edición ─────────────────────────

    def _build_tab_edicion(self, parent):
        toolbar = tk.Frame(parent, bg=BG, pady=16)
        toolbar.pack(fill="x", padx=10)

        info_lbl = tk.Label(toolbar, text="Panel de Administración y Edición de Productos", bg=BG, fg=TEXT_DIM, font=("Segoe UI", 11, "italic"))
        info_lbl.pack(side="top", anchor="w", pady=(0, 10))
        
        btn_frame = tk.Frame(toolbar, bg=BG)
        btn_frame.pack(fill="x")

        # Botones que se bloquean con el lock de edición
        b_nuevo      = styled_btn(btn_frame, "+ Nuevo Producto", self._nuevo_producto,    color=SUCCESS, width=16)
        b_editar     = styled_btn(btn_frame, "✎ Editar Prod.",    self._editar_producto,   color=ACCENT,  width=16)
        b_aumento    = styled_btn(btn_frame, "📈 Precios",        self._abrir_aumento,     color=WARNING, width=16)
        b_redondeo_up = styled_btn(btn_frame, "⬆ Redondear", self._redondear_up, color="#a35fcc", width=12)
        b_redondeo_dn = styled_btn(btn_frame, "⬇ Redondear", self._redondear_dn, color="#a35fcc", width=12)
        b_eliminar   = styled_btn(btn_frame, "🗑 Eliminar",        self._eliminar_producto, color="#8b3a3a", width=16)
        # 1. Empacar PRIMERO los botones de la derecha para evitar que se recorten por falta de espacio
        # (se empacan en orden inverso porque pack(side="right") los apila hacia la izquierda)
        
        # Botón Actualizar
        b_actualizar = styled_btn(btn_frame, "🔄 Actualizar", self.refresh_tab_edicion, color=BG3)
        b_actualizar.pack(side="right", padx=6)

        # Menú desplegable Exportar / Importar
        mb_datos = tk.Menubutton(
            btn_frame, text="📁 Datos ▾",
            bg=BG3, fg=TEXT, relief="flat",
            font=("Segoe UI", 9, "bold"),
            padx=10, pady=6, cursor="hand2"
        )
        menu_datos = tk.Menu(mb_datos, tearoff=0, bg=BG3, fg=TEXT,
                             activebackground=ACCENT, activeforeground=TEXT,
                             font=("Segoe UI", 9))
        menu_datos.add_command(label="⬇  Exportar XLSX", command=self._exportar_excel)
        menu_datos.add_command(label="⬆  Importar archivo", command=self._importar_datos)
        menu_datos.add_separator()
        menu_datos.add_command(label="💰  Reporte de Capital", command=self._abrir_reporte_capital)
        menu_datos.add_separator()
        menu_datos.add_command(label="📦  Crear Copia de Seguridad", command=self._backup_db)
        menu_datos.add_separator()
        menu_datos.add_command(label="📄  Convertir PDF a Excel", command=self._abrir_ilovepdf)
        menu_datos.add_command(label="📋  Generar Lista de Precios PDF", command=self._abrir_generar_lista_pdf)
        menu_datos.add_command(label="📊  Comparador de Listas", command=self._abrir_comparador_listas)
        menu_datos.add_command(label="📈  Dashboard de Ventas", command=self._abrir_dashboard_ventas)
        menu_datos.add_command(label="📝  Hoja de Cálculo", command=self._abrir_hoja_calculo)
        mb_datos.config(menu=menu_datos)
        mb_datos.pack(side="right", padx=6)
        
        # Botón Limpiar Selección explícito
        b_limpiar_ed = styled_btn(btn_frame, "🧹 Limpiar selección", self._deseleccionar_todos_edicion, color=BG3)
        b_limpiar_ed.pack(side="right", padx=6)

        # 2. Empacar botones de la izquierda
        b_nuevo.pack(side="left", padx=6)
        b_editar.pack(side="left", padx=6)
        b_aumento.pack(side="left", padx=6)
        b_redondeo_up.pack(side="left", padx=2)
        b_redondeo_dn.pack(side="left", padx=2)

        tk.Frame(btn_frame, width=10, bg=BG).pack(side="left")  # Separador
        b_entrada = styled_btn(btn_frame, "▲ Entrada", self._entrada_stock, color="#4fa882", width=12)
        b_entrada.pack(side="left", padx=6)
        self._btn_entrada = b_entrada
        
        b_stock_rapido = styled_btn(btn_frame, "⚡ Stock Rápido", self._toggle_modo_stock_rapido, color="#2563eb", width=14)
        b_stock_rapido.pack(side="left", padx=6)
        self._btn_stock_rapido = b_stock_rapido
        
        tk.Frame(btn_frame, width=10, bg=BG).pack(side="left")  # Separador
        b_eliminar.pack(side="left", padx=6)
        
        # ── Opciones Previas y Alertas ──────────────
        info_foot = tk.Label(parent, text="⚠ Selecciona los productos que deseas afectar desde la lista de abajo antes de utilizar estas herramientas.", bg=BG, fg=WARNING, font=("Segoe UI", 9))
        info_foot.pack(side="top", anchor="w", padx=10, pady=(4, 0))

        # Inicializamos `self.tree_ed` aquí arriba para evitar AttributeError en los callbacks posteriores
        tree_frame_ed = tk.Frame(parent, bg=BG)
        # Se empacará después del buscador
        self.tree_ed = ttk.Treeview(tree_frame_ed, columns=("sel", "codigo", "nombre", "marca", "categoria", "stock", "costo", "precio"), show="headings", selectmode="extended")

        # Búsqueda + filtros de categoría / marca para Edición
        search_frame_ed = tk.Frame(parent, bg=BG)
        search_frame_ed.pack(fill="x", padx=10, pady=(0, 6))

        tk.Label(search_frame_ed, text="🔍", bg=BG, fg=TEXT_DIM,
                 font=("Segoe UI", 11)).pack(side="left", padx=(0, 4))
        self.v_buscar_edicion = tk.StringVar()
        self.v_buscar_edicion.trace_add("write", lambda *_: self.refresh_tab_edicion())
        self.entry_buscar_edicion = entry(search_frame_ed, textvariable=self.v_buscar_edicion, width=28)
        self.entry_buscar_edicion.pack(side="left")
        self.v_buscar_edicion.set(" Buscar (F1)...")
        self.entry_buscar_edicion.config(fg='gray')

        def _on_focus_in_buscar_ed(e):
            if self.v_buscar_edicion.get() == " Buscar (F1)...":
                self.v_buscar_edicion.set("")
                self.entry_buscar_edicion.config(fg=TEXT)
        def _on_focus_out_buscar_ed(e):
            if not self.v_buscar_edicion.get():
                self.v_buscar_edicion.set(" Buscar (F1)...")
                self.entry_buscar_edicion.config(fg='gray')

        self.entry_buscar_edicion.bind("<FocusIn>", _on_focus_in_buscar_ed)
        self.entry_buscar_edicion.bind("<FocusOut>", _on_focus_out_buscar_ed)

        # Categoría
        tk.Label(search_frame_ed, text="Categoría:", bg=BG, fg=TEXT, font=("Segoe UI", 9)).pack(side="left", padx=(16, 4))
        self.v_cat_filtro_ed = tk.StringVar(value="Todas")
        self.cmb_cat_ed = ttk.Combobox(search_frame_ed, textvariable=self.v_cat_filtro_ed, state="readonly", width=16)
        self.cmb_cat_ed.pack(side="left", padx=0)
        self.v_cat_filtro_ed.trace_add("write", lambda *_: self.refresh_tab_edicion())

        # Marca
        tk.Label(search_frame_ed, text="Marca:", bg=BG, fg=TEXT, font=("Segoe UI", 9)).pack(side="left", padx=(16, 4))
        self.v_marca_filtro_ed = tk.StringVar(value="Todas")
        self.cmb_marca_ed = ttk.Combobox(search_frame_ed, textvariable=self.v_marca_filtro_ed, state="readonly", width=16)
        self.cmb_marca_ed.pack(side="left", padx=0)
        self.v_marca_filtro_ed.trace_add("write", lambda *_: self.refresh_tab_edicion())

        # Atajos helper text (sutil en el fondo)
        lbl_atajos_ed = tk.Label(parent, text="Atajos: [F1] Foco Buscador  •  [F2] Seleccionar  •  [F5] Actualizar  •  [↑/↓] Navegar  •  [Enter] Editar Item / Aumento Masivo", bg=BG, fg=TEXT_DIM, font=("Segoe UI", 8))
        lbl_atajos_ed.pack(side="bottom", anchor="w", padx=14, pady=(0, 6))

        # ── Lista de productos (duplicado editable) ──────────────
        tree_frame_ed.pack(fill="both", expand=True, padx=10, pady=(4, 8))

        cols_ed = ("sel", "codigo", "nombre", "marca", "categoria", "stock", "costo", "precio")
        hdrs_ed = ("[F2]", "Código", "Nombre", "Marca", "Categoría", "Stock", "Costo", "Precio Venta")
        widths_ed   = (35, 90, 200, 110, 110, 70, 80, 80)
        for col, hdr, w in zip(cols_ed, hdrs_ed, widths_ed):
            self.tree_ed.heading(col, text=hdr)
            self.tree_ed.column(col, width=w, anchor="w" if col == "nombre" else "center")

        self.tree_ed.tag_configure("bajo",  background="#4a2020")
        self.tree_ed.tag_configure("cero",  background="#3a1010")
        self.tree_ed.tag_configure("alt",   background=ROW_ALT)
        self.tree_ed.tag_configure("seleccionado", foreground=SUCCESS)

        sb_ed = ttk.Scrollbar(tree_frame_ed, orient="vertical", command=self.tree_ed.yview)
        self.tree_ed.configure(yscrollcommand=sb_ed.set)
        self.tree_ed.pack(side="left", fill="both", expand=True)
        sb_ed.pack(side="right", fill="y")

        self.tree_ed.bind("<Double-1>", lambda e: self._editar_producto_desde_tree_ed() if self.tree_ed.identify_region(e.x, e.y) == "cell" else None)
        
        def _treeview_sort_column_ed(col_id, rev, is_sel_col=False):
            items = [(self.tree_ed.set(k, col_id), k) for k in self.tree_ed.get_children("")]
            if is_sel_col:
                items.sort(key=lambda t: 0 if t[0] == "[F2]" else 1, reverse=rev)
            else:
                try:
                    items.sort(key=lambda t: float(t[0].replace("$","")) if t[0].replace("$","").replace(".","").isdigit() else t[0].lower(), reverse=rev)
                except Exception:
                    items.sort(key=lambda t: t[0].lower(), reverse=rev)
                    
            for idx, (_, k) in enumerate(items):
                self.tree_ed.move(k, "", idx)
            self.tree_ed.heading(col_id, command=lambda _c=col_id: _treeview_sort_column_ed(_c, not rev, is_sel_col=is_sel_col))
            
        for col in cols_ed:
            if col == "sel":
                self.tree_ed.heading(col, command=lambda _c=col: _treeview_sort_column_ed(_c, False, is_sel_col=True))
            else:
                self.tree_ed.heading(col, command=lambda _c=col: _treeview_sort_column_ed(_c, False))

        self.tree_ed.bind("<ButtonRelease-1>", self._toggle_sel_prod)
        self.tree_ed.bind("<ButtonRelease-1>", self._iniciar_edicion_stock_celda, add="+")
        self.tree_ed.bind("<Shift-ButtonRelease-1>", self._shift_toggle_sel_prod)
        self.tree_ed.bind("<B1-Motion>", self._drag_sel_prod)
        self.tree_ed.bind("<ButtonPress-1>", self._start_drag_sel_prod)
        
        # Panel Flotante de Confirmación (Oculto por defecto)
        self.frame_flotante_stock = tk.Frame(tree_frame_ed, bg=BG3, highlightbackground=SUCCESS, highlightthickness=2, padx=15, pady=10)
        tk.Label(self.frame_flotante_stock, text="Modo Edición Rápida de Stock", bg=BG3, fg=TEXT, font=("Segoe UI", 10, "bold")).pack(side="top", pady=(0, 10))
        
        btn_box_flotante = tk.Frame(self.frame_flotante_stock, bg=BG3)
        btn_box_flotante.pack(side="bottom")
        
        b_conf = styled_btn(btn_box_flotante, "✔ Confirmar Stock", self._confirmar_stock_masivo, color=SUCCESS)
        b_conf.pack(side="left", padx=5)
        
        b_canc = styled_btn(btn_box_flotante, "✖ Cancelar", self._cancelar_stock_masivo, color=DANGER)
        b_canc.pack(side="left", padx=5)

        def _on_tree_ed_enter(event):
            sel = self.tree_ed.selection()
            if not sel: return "break"
            if len(self._seleccionados) > 1:
                self._abrir_aumento()
            else:
                self._editar_producto()
            return "break"
            
        def _on_tree_ed_f2(event):
            sel = self.tree_ed.selection()
            if not sel: return "break"
            iid = sel[0]
            if iid in self._seleccionados:
                self._seleccionados.remove(iid)
            else:
                self._seleccionados.add(iid)
                
            if hasattr(self, 'refresh_tab_edicion'):
                self.refresh_tab_edicion()
            # Mantener foco
            self.tree_ed.focus_set()
            self.tree_ed.focus(iid)
            self.tree_ed.selection_set(iid)
            return "break"
        
        self.tree_ed.bind("<Return>", _on_tree_ed_enter)
        self.tree_ed.bind("<F2>", _on_tree_ed_f2)

        # Guardar referencias para controlar con el lock
        self._btns_edicion = [b_nuevo, b_editar, b_aumento, b_redondeo_up, b_redondeo_dn, b_eliminar, mb_datos, b_entrada]
        self._aplicar_estado_edicion()

    # ── MÉTODOS MODO STOCK RÁPIDO ─────────────────────────

    def _toggle_modo_stock_rapido(self):
        if not self.v_edicion.get():
            messagebox.showwarning("Atención", "Desbloquea el modo administrador (Candado verde) para utilizar la edición rápida.", parent=self)
            return

        self._modo_stock_rapido = not self._modo_stock_rapido
        
        if self._modo_stock_rapido:
            # Activar Modo
            self._btn_stock_rapido.config(text="Volver Normal", bg=WARNING, fg=BG, activebackground="#e5a322", activeforeground=BG)
            self.tree_ed.heading("stock", text="⚡ [ EN EDICIÓN ]")
            self.frame_flotante_stock.place(relx=0.5, rely=0.9, anchor="center")
            self._cambios_stock_tmp = {}
            self._undo_stock_stack = []
            
            # Desactivar botones momentaneamente
            for btn in self._btns_edicion:
                try: btn.config(state="disabled")
                except Exception as e: logging.exception("Error silencioso genérico capturado:")
                
            messagebox.showinfo("Stock Rápido Activado", "Haz clic en cualquier celda de la columna de STOCK para editarla.\nTus cambios no se guardarán en DB hasta que des a 'Confirmar'.\n\n- Presiona 'Tab' después de editar para pasar al producto de abajo.\n- Usa 'Ctrl+Z' para deshacer cambios individuales temporalmente.", parent=self)
        else:
            # Desactivar Modo Seguro
            if self._cambios_stock_tmp:
                if messagebox.askyesno("Cambios pendientes", "Tienes modificaciones de stock sin confirmar.\n¿Deseas aplicarlas ahora antes de salir?\n\n(Sí = Confirmar cambios, No = Descartar y salir)", parent=self):
                    self._confirmar_stock_masivo()
                    return
                else:
                    self._cancelar_stock_masivo()  # descarta los visuales
                    return # el cancel ya hace el toggle
                    
            self._btn_stock_rapido.config(text="⚡ Stock Rápido", bg="#2563eb", fg="white", activebackground="#1d4ed8", activeforeground="white")
            self.tree_ed.heading("stock", text="Stock")
            self.frame_flotante_stock.place_forget()
            
            self._aplicar_estado_edicion() # restaurar botones

    def _iniciar_edicion_stock_celda(self, event):
        if not self._modo_stock_rapido: return

        # Identificar qué celda se clickeó
        region = self.tree_ed.identify_region(event.x, event.y)
        if region != "cell": return
        
        col_id = self.tree_ed.identify_column(event.x)
        if col_id != "#6": return # #6 es la columna de Stock
        
        iid = self.tree_ed.identify_row(event.y)
        if not iid: return
        
        self._montar_entry_stock(iid)

    def _montar_entry_stock(self, iid, delay=True):
        # Evitar doble edición en la misma o diferente si no ha cerrado
        if hasattr(self, '_stock_entry_active') and self._stock_entry_active:
            try:
                self._stock_entry_active.destroy()
            except Exception as e: logging.exception("Error silencioso capturado:")
            self._stock_entry_active = None

        if delay:
            # Diferir la creación del Entry para que el Treeview termine de procesar el click
            self.after(30, lambda: self._crear_entry_stock(iid))
        else:
            self.tree_ed.update_idletasks()
            self._crear_entry_stock(iid)

    def _crear_entry_stock(self, iid):
        prod_id = iid # El iid de la fila ES el id de la base de datos
        
        prod_data = db.get_producto_by_id(prod_id)
        if prod_data:
            val_limpio = str(prod_data.get("stock", 0))
            if val_limpio.endswith(".0"):
                val_limpio = val_limpio[:-2]
        else:
            val_limpio = "0"
        
        bbox = self.tree_ed.bbox(iid, "#6")
        if not bbox: return
        x, y, w, h = bbox
        
        entry = ttk.Entry(self.tree_ed, font=("Segoe UI", 10), justify="center")
        entry.place(x=x, y=y, width=w, height=h)
        self._stock_entry_active = entry
        entry.delete(0, tk.END)
        entry.insert(0, val_limpio)
        entry.select_range(0, tk.END)
        entry.focus_force()

        def validate_and_save(e=None, nav=None):
            if not getattr(self, '_stock_entry_active', None) or not entry.winfo_exists(): return "break"
            
            nuevo_texto = entry.get().replace(' Kg', '').replace(' kg', '').strip()
            try:
                if nuevo_texto == "": nuevo_val = 0.0
                else: nuevo_val = float(nuevo_texto)
                if nuevo_val < 0: raise ValueError
            except ValueError:
                messagebox.showerror("Error", "Debes ingresar un número válido de stock (ej. 10 o 5.5).", parent=self)
                entry.focus_force()
                entry.select_range(0, tk.END)
                return "break"
            
            try:
                prod = db.obtener_producto_por_id(prod_id)
                if prod:
                    stock_original_db = float(prod.get("stock", 0))
                else:
                    stock_original_db = 0.0
            except Exception: stock_original_db = 0.0

            stock_previo_visual = val_limpio
            self._undo_stock_stack.append((iid, prod_id, stock_previo_visual))
            
            if stock_original_db == nuevo_val:
                if prod_id in self._cambios_stock_tmp:
                    del self._cambios_stock_tmp[prod_id]
            else:
                self._cambios_stock_tmp[prod_id] = nuevo_val
            
            format_str = f"{nuevo_val:g} kg" if val_actual.endswith("kg") else f"{nuevo_val:g}"
            self.tree_ed.set(iid, "stock", format_str)
            self.tree_ed.item(iid, tags=("seleccionado",))

            try: entry.destroy()
            except Exception as e: logging.exception("Error silencioso capturado:")
            self._stock_entry_active = None
            
            # Desvincular el evento global de clic si existiera
            if hasattr(self, '_stock_click_bg_bind'):
                try: self.tree_ed.unbind("<Button-1>", self._stock_click_bg_bind)
                except Exception as e: logging.exception("Error silencioso capturado:")
                delattr(self, '_stock_click_bg_bind')

            if nav:
                siguiente = None
                if nav in ('tab', 'down'):
                    siguiente = self.tree_ed.next(iid)
                elif nav == 'up':
                    siguiente = self.tree_ed.prev(iid)
                
                if siguiente:
                    self.tree_ed.see(siguiente)
                    self.tree_ed.selection_set(siguiente)
                    self._montar_entry_stock(siguiente, delay=False)
                else:
                    self.tree_ed.focus_force()
                    self.tree_ed.selection_set(iid)
                    
            return "break"

        def cancel(e=None):
            try: entry.destroy()
            except Exception as e: logging.exception("Error silencioso capturado:")
            self._stock_entry_active = None
            if hasattr(self, '_stock_click_bg_bind'):
                try: self.tree_ed.unbind("<Button-1>", self._stock_click_bg_bind)
                except Exception as e: logging.exception("Error silencioso capturado:")
                delattr(self, '_stock_click_bg_bind')

        def on_tree_click_bg(e):
            # Si hace clic fuera del entry, guardar y cerrar
            if entry.winfo_exists():
                validate_and_save(e, nav=None)
        def validate_wrapper(e, nav_action):
            validate_and_save(e, nav=nav_action)
            return "break"
            
        entry.bind("<Return>", lambda e: validate_wrapper(e, None))
        entry.bind("<Tab>", lambda e: validate_wrapper(e, 'tab'))
        entry.bind("<Down>", lambda e: validate_wrapper(e, 'down'))
        entry.bind("<Up>", lambda e: validate_wrapper(e, 'up'))
        entry.bind("<Escape>", cancel)
        
        # Detectar clic en otra parte de la tabla para auto-guardar
        self._stock_click_bg_bind = self.tree_ed.bind("<Button-1>", on_tree_click_bg, add="+")

    def _undo_stock_rapido_action(self, event=None):
        if not getattr(self, '_modo_stock_rapido', False) or not getattr(self, '_undo_stock_stack', []): return "break"
        
        iid, prod_id, stock_previo_visual = self._undo_stock_stack.pop()
        
        try:
            prod = db.obtener_producto_por_id(prod_id)
            if prod:
                stock_original_db = float(prod.get("stock", 0))
                es_peso = prod.get("se_vende_por_peso", 0)
            else:
                return "break"
        except Exception: 
            return "break"

        num_prev = float(stock_previo_visual) if stock_previo_visual else 0.0

        if stock_original_db == num_prev:
            if prod_id in self._cambios_stock_tmp:
                del self._cambios_stock_tmp[prod_id]
        else:
            self._cambios_stock_tmp[prod_id] = num_prev

        # Restore visual
        format_str = f"{num_prev:g} kg" if es_peso else f"{num_prev:g}"
        try:
            self.tree_ed.set(iid, "stock", format_str)
            i_idx = self.tree_ed.index(iid)
            tag = "alt" if i_idx % 2 != 0 else ""
            self.tree_ed.item(iid, tags=(tag,))
        except Exception as e: logging.exception("Error silencioso capturado:")
        return "break"

    def _confirmar_stock_masivo(self):
        # Si un campo está activo, forzar guardado primero
        if getattr(self, '_stock_entry_active', None) and self._stock_entry_active.winfo_exists():
            # Forzamos un evento Return sintético al entry para grabarlo
            self._stock_entry_active.event_generate("<Return>")
            
        if not self._cambios_stock_tmp:
            messagebox.showinfo("Aviso", "No se detectaron cambios de stock para guardar.", parent=self)
            self._toggle_modo_stock_rapido()
            return
        
        cambios = 0
        from datetime import datetime
        ahora = datetime.now().isoformat()
        
        for p_id, q_nueva in self._cambios_stock_tmp.items():
            prod = db.get_producto_by_id(p_id)
            if not prod: continue
            
            q_vieja = prod["stock"]
            dif = q_nueva - q_vieja
            if dif == 0: continue
            
            tipo_mov = "entrada" if dif > 0 else "salida"
            dif_abs = abs(dif)

            # registrar_movimiento ya actualiza el stock automáticamente
            db.registrar_movimiento(
                producto_id=p_id,
                tipo=tipo_mov,
                cantidad=dif_abs,
                nota="Ajuste de Stock Rápido (Lote)",
                forzar=True
            )
            
            cambios += 1
        
        self.refresh_productos()
        self.refresh_historial()
        self.refresh_alertas()
        self.refresh_tab_edicion()
        
        messagebox.showinfo("Proceso Terminado", f"Se han registrado {cambios} correcciones de stock en la base de datos.", parent=self)
        
        self._cambios_stock_tmp.clear()
        self._toggle_modo_stock_rapido() 

    def _cancelar_stock_masivo(self):
        if hasattr(self, '_stock_click_bg_bind'):
            try: self.tree_ed.unbind("<Button-1>", self._stock_click_bg_bind)
            except Exception as e: logging.exception("Error silencioso capturado:")
            delattr(self, '_stock_click_bg_bind')
            
        if self._cambios_stock_tmp:
            if not messagebox.askyesno("Cancelar", "Esto descartará todos los cambios temporales de stock a cómo estaban registrados antes.\n\n¿Estás seguro?", parent=self):
                return
        
        # Eliminar el entry activo si existe
        if getattr(self, '_stock_entry_active', None):
            try:
                self._stock_entry_active.destroy()
            except Exception as e: logging.exception("Error silencioso capturado:")
            self._stock_entry_active = None

        self._modo_stock_rapido = False
        self._btn_stock_rapido.config(text="⚡ Stock Rápido", bg="#2563eb", fg="white", activebackground="#1d4ed8", activeforeground="white")
        self.tree_ed.heading("stock", text="Stock")
        self.frame_flotante_stock.place_forget()
        self._cambios_stock_tmp.clear()
        self.refresh_tab_edicion()
        self._aplicar_estado_edicion()

    def _backup_db(self):
        import shutil, os, datetime
        
        fecha = datetime.datetime.now().strftime("%Y-%m-%d_%H%M%S")
        nombre_def = f"stock_backup_{fecha}.db"
        
        filepath = filedialog.asksaveasfilename(
            defaultextension=".db",
            initialfile=nombre_def,
            title="Guardar copia de seguridad",
            filetypes=[("Base de Datos SQLite", "*.db"), ("Todos los archivos", "*.*")]
        )
        
        if filepath:
            try:
                shutil.copy2("stock.db", filepath)
                messagebox.showinfo("Backup Exitoso", f"Copia de seguridad guardada correctamente en:\n{filepath}", parent=self)
            except Exception as e:
                messagebox.showerror("Error de Backup", f"No se pudo crear la copia de seguridad:\n{e}", parent=self)

    def _abrir_comparador_listas(self):
        import webbrowser
        import sys
        from pathlib import Path
        
        if getattr(sys, 'frozen', False):
            # Si corre desde el .exe (usamos sys._MEIPASS que es donde PyInstaller extrae los datos temporales)
            base_dir = Path(sys._MEIPASS)
        else:
            base_dir = Path(__file__).parent
            
        ruta = base_dir / "comparador_listas.html"
        
        if ruta.exists():
            webbrowser.open(ruta.as_uri())
        else:
            # Fallback en caso de que lo abran como .exe pero no esté en MEIPASS sino al lado
            ruta_alt = Path(sys.executable).parent / "comparador_listas.html"
            if ruta_alt.exists():
                webbrowser.open(ruta_alt.as_uri())
            else:
                messagebox.showerror("Archivo no encontrado", f"No se encontró el comparador en:\n{ruta}\nni en:\n{ruta_alt}")
            
    def _abrir_hoja_calculo(self):
        import webbrowser
        webbrowser.open("https://docs.google.com/spreadsheets/u/0/")

    
    def _abrir_ilovepdf(self):
        ILovePdfDialog(self)

    def _abrir_generar_lista_pdf(self):
        GenerarListaPdfDialog(self)

    def _abrir_dashboard_ventas(self):
        import webbrowser
        import sys
        from pathlib import Path
        
        if getattr(sys, 'frozen', False):
            base_dir = Path(sys._MEIPASS)
        else:
            base_dir = Path(__file__).parent
            
        ruta = base_dir / "dashboard_ventas.html"
        
        if ruta.exists():
            webbrowser.open(ruta.as_uri())
        else:
            ruta_alt = Path(sys.executable).parent / "dashboard_ventas.html"
            if ruta_alt.exists():
                webbrowser.open(ruta_alt.as_uri())
            else:
                messagebox.showerror("Error", "No se encontró el archivo dashboard_ventas.html")

    # ── Pestaña Productos ─────────────────────────

    def _build_tab_productos(self, parent):
        # Barra POS: solo el botón Carrito
        toolbar = tk.Frame(parent, bg=BG, pady=8)
        toolbar.pack(fill="x", padx=10)

        # Botón carrito gigante
        self.btn_carrito = tk.Button(
            toolbar, text="🛒 Carrito (F3)", command=self._abrir_carrito,
            bg=ACCENT2, fg=TEXT, relief="flat",
            padx=20, pady=8, font=("Segoe UI", 12, "bold"),
            activebackground=ACCENT, activeforeground=TEXT, cursor="hand2"
        )
        self.btn_carrito.pack(side="right", padx=4)
        
        # Botón Actualizar explícito gui
        styled_btn(toolbar, "🔄 Actualizar (F5)", self._refresh_all_tabs, color=BG3).pack(side="right", padx=6)
        
        # Botón Limpiar Selección explícito
        styled_btn(toolbar, "🧹 Limpiar selección", self._deseleccionar_todos, color=BG3).pack(side="right", padx=6)

        # Búsqueda + filtro de categoría
        search_frame = tk.Frame(parent, bg=BG)
        search_frame.pack(fill="x", padx=10, pady=(0, 6))

        tk.Label(search_frame, text="🔍", bg=BG, fg=TEXT_DIM,
                 font=("Segoe UI", 11)).pack(side="left", padx=(0, 4))
                 
        self.v_buscar = tk.StringVar(value=" Buscar (F1)...")
        self.entry_buscar = entry(search_frame, textvariable=self.v_buscar, width=28)
        self.entry_buscar.pack(side="left")
        self.entry_buscar.config(fg='gray')

        def _on_buscar_focus_in(e):
            if self.v_buscar.get() == " Buscar (F1)...":
                self.v_buscar.set("")
                self.entry_buscar.config(fg=TEXT)
                
        def _on_buscar_focus_out(e):
            if not self.v_buscar.get():
                self.v_buscar.set(" Buscar (F1)...")
                self.entry_buscar.config(fg='gray')
                
        self.entry_buscar.bind("<FocusIn>", _on_buscar_focus_in)
        self.entry_buscar.bind("<FocusOut>", _on_buscar_focus_out)

        tk.Label(search_frame, text=" Categoría:", bg=BG, fg=TEXT_DIM,
                 font=("Segoe UI", 8)).pack(side="left", padx=(8, 2))
                 
        self.v_cat_filtro = tk.StringVar(value="Cualquiera")
        self.cmb_cat = ttk.Combobox(
            search_frame, textvariable=self.v_cat_filtro,
            values=["Cualquiera"], width=13, state="readonly"
        )
        self.cmb_cat.pack(side="left")

        tk.Label(search_frame, text=" Marca:", bg=BG, fg=TEXT_DIM,
                 font=("Segoe UI", 8)).pack(side="left", padx=(8, 2))
                 
        self.v_marca_filtro = tk.StringVar(value="Cualquiera")
        self.cmb_marca = ttk.Combobox(
            search_frame, textvariable=self.v_marca_filtro,
            values=["Cualquiera"], width=13, state="readonly"
        )
        self.cmb_marca.pack(side="left", padx=(0, 4))

        # Add traces AFTER all vars are ready
        self.v_buscar.trace_add("write", lambda *_: self.refresh_productos())
        self.v_cat_filtro.trace_add("write", lambda *_: self.refresh_productos())
        self.v_marca_filtro.trace_add("write", lambda *_: self.refresh_productos())

        # Tabla
        cols = ("sel", "codigo", "nombre", "categoria", "marca", "stock", "minimo", "precio")
        headers = ("[F2]", "Código", "Nombre", "Categoría", "Marca", "Stock", "Mínimo", "Precio")

        # Atajos helper text (sutil en el fondo)
        lbl_atajos = tk.Label(parent, text="Atajos: [F1] Foco Buscador  •  [F2] Seleccionar  •  [F5] Actualizar  •  [↑/↓] Navegar  •  [Enter / Espacio] Carrito  •  [F3] Abrir Carrito  •  [F6] Notas", bg=BG, fg=TEXT_DIM, font=("Segoe UI", 8))
        lbl_atajos.pack(side="bottom", anchor="w", padx=14, pady=(0, 6))

        tree_frame = tk.Frame(parent, bg=BG)
        tree_frame.pack(fill="both", expand=True, padx=10, pady=(0, 6))

        self.tree = ttk.Treeview(tree_frame, columns=cols, show="headings", selectmode="extended")
        widths = (35, 100, 220, 130, 130, 70, 70, 85)
        for col, hdr, w in zip(cols, headers, widths):
            self.tree.heading(col, text=hdr,
                              command=lambda c=col: self._sort_column(c))
            self.tree.column(col, width=w, anchor="center" if col not in ("nombre","categoria","marca") else "w")

        self.tree.tag_configure("stock_bajo", foreground=WARNING)
        self.tree.tag_configure("sin_stock",  foreground=DANGER)
        self.tree.tag_configure("alt",        background=ROW_ALT)
        self.tree.tag_configure("alt_bajo",   background=ROW_ALT, foreground=WARNING)
        self.tree.tag_configure("seleccionado", foreground=SUCCESS)
        self.tree.tag_configure("con_nota", background="#e1d8ea")
        
        sb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=sb.set)

        self.tree.pack(side="left", fill="both", expand=True)
        sb.pack(side="right", fill="y")

        self.tree.bind("<Double-1>", self._on_tree_double_click)
        self.tree.bind("<ButtonRelease-1>", self._toggle_sel_prod)
        self.tree.bind("<Shift-ButtonRelease-1>", self._shift_toggle_sel_prod)
        self.tree.bind("<B1-Motion>", self._drag_sel_prod)
        self.tree.bind("<ButtonPress-1>", self._start_drag_sel_prod)

        self._sort_col = None
        self._sort_rev = False

        # Desactivar la navegación con tabulador excepto cuando hay un Entry de edición de stock activo
        def _maybe_block_tab(event):
            if getattr(self, '_modo_stock_rapido', False) and getattr(self, '_stock_entry_active', None):
                return None  # Dejar pasar el Tab al Entry
            return "break"
        self.bind_all("<Tab>", _maybe_block_tab)
        self.bind_all("<Shift-Tab>", _maybe_block_tab)

        # --- Atajos de teclado globales POS ---
        def get_active_tab():
            try: return self.notebook.index(self.notebook.select())
            except: return -1

        def _focus_search_f1(e):
            idx = get_active_tab()
            if idx == 0: self.entry_buscar.focus_set()
            elif idx == 2: self.entry_buscar_alertas.focus_set()
            elif idx == 3: self.entry_buscar_edicion.focus_set()

        self.bind_all("<F1>", _focus_search_f1)
        self.bind("<F6>", lambda e: self._on_f6_pressed() if get_active_tab() == 0 else None)
        
        def _focus_search_on_type(event):
            idx = get_active_tab()
            if idx not in (0, 2, 3): return
            if not isinstance(event.widget, (tk.Entry, ttk.Combobox, tk.Text)):
                # Ignore control combos or non-printable chars
                if getattr(event, 'char', None) and event.char.isprintable() and not (event.state & 4) and event.keysym not in ("F1", "F2", "F3", "F4", "F5", "F6", "F7", "F8", "F9", "F10", "F11", "F12"):
                    target_entry = None
                    target_var = None
                    if idx == 0: 
                        target_entry = self.entry_buscar
                        target_var = self.v_buscar
                    elif idx == 2:
                        target_entry = self.entry_buscar_alertas
                        target_var = self.v_buscar_alertas
                    elif idx == 3:
                        target_entry = self.entry_buscar_edicion
                        target_var = self.v_buscar_edicion
                        
                    if target_entry and target_var:
                        target_entry.focus_set()
                        if target_var.get().startswith(" Buscar"):
                            target_var.set("")
                            target_entry.config(fg=TEXT)
                        target_entry.insert(tk.END, event.char)
                        return "break"
        self.bind("<Key>", _focus_search_on_type)

        def _tree_select_first(event):
            idx = get_active_tab()
            if idx == 0:
                sel = self.tree.selection()
                if not sel:
                    items = self.tree.get_children()
                    if items:
                        self.tree.selection_set(items[0])
                        self.tree.focus(items[0])
                        self.tree.focus_set()
                        return "break"
                else:
                    if not isinstance(event.widget, ttk.Treeview):
                        self.tree.focus_set()
                        return "break"
            elif idx == 3:
                sel = self.tree_ed.selection()
                if not sel:
                    items = self.tree_ed.get_children()
                    if items:
                        self.tree_ed.selection_set(items[0])
                        self.tree_ed.focus(items[0])
                        self.tree_ed.focus_set()
                        return "break"
                else:
                    if not isinstance(event.widget, ttk.Treeview):
                        self.tree_ed.focus_set()
                        return "break"
        self.bind("<Down>", _tree_select_first)
        self.bind("<Up>", _tree_select_first)

        def _add_to_cart_kbd(event):
            idx = get_active_tab()
            if idx != 0: return "break"
            sel = self.tree.selection()
            if sel:
                self._on_tree_enter(event)
                return "break"
        
        self.tree.bind("<Return>", _add_to_cart_kbd)
        self.tree.bind("<space>", _add_to_cart_kbd)
        
        def _on_tree_productos_f2(event):
            sel = self.tree.selection()
            if not sel: return "break"
            iid = sel[0]
            
            try: y_pos = self.tree.yview()
            except: y_pos = None

            # Encontrar el elemento siguiente (para dejar el foco y selección ahí)
            next_iid = self.tree.next(iid)
            if not next_iid:
                next_iid = self.tree.prev(iid) # Si es el último, saltar al anterior

            if iid in getattr(self, "_seleccionados_pos", set()):
                self._seleccionados_pos.remove(iid)
            else:
                self._seleccionados_pos.add(iid)

            self.refresh_productos()
            
            self.tree.focus_set()
            if next_iid and self.tree.exists(next_iid):
                self.tree.focus(next_iid)
                self.tree.selection_set(next_iid)
            elif self.tree.exists(iid):
                # Fallback por si la lista se vació/cambió tanto que no hay next ni prev
                self.tree.focus(iid)
                self.tree.selection_set(iid)
            
            if y_pos:
                self.tree.after(10, lambda: self.tree.yview_moveto(y_pos[0]))
                
            return "break"
            
        self.tree.bind("<F2>", _on_tree_productos_f2)
        
        self.bind_all("<F3>", lambda e: self._abrir_carrito() if get_active_tab() == 0 else None)

    # ── Pestaña Historial ─────────────────────────

    def _build_tab_historial(self, parent):
        toolbar = tk.Frame(parent, bg=BG, pady=8)
        toolbar.pack(fill="x", padx=10)

        styled_btn(toolbar, "🔄 Actualizar", self.refresh_historial, color=BG3).pack(side="left", padx=4)
        # Selector: Modo de agrupación
        tk.Label(toolbar, text="Agrupar por:", bg=BG, fg=TEXT, font=("Segoe UI", 9)).pack(side="left", padx=(16, 4))
        self.v_agrupar = tk.StringVar(value="Sin agrupar")
        cmb_agrupar = ttk.Combobox(
            toolbar, textvariable=self.v_agrupar,
            values=["Sin agrupar", "Producto", "Orden de Venta"],
            state="readonly", width=16
        )
        cmb_agrupar.pack(side="left", padx=0)
        self.v_agrupar.trace_add("write", lambda *_: self.refresh_historial())

        self.btn_export_ventas = styled_btn(toolbar, "📤 Exportar ventas", self._exportar_ventas, color="#2a5f3a")
        self.btn_export_ventas.pack(side="right", padx=4)
        self.v_agrupar.trace_add("write", lambda *_: self.refresh_historial())

        # Filtros por columna
        filter_frame = tk.Frame(parent, bg=BG)
        filter_frame.pack(fill="x", padx=10, pady=(0, 4))

        tk.Label(filter_frame, text="Filtros:", bg=BG, fg=TEXT_DIM,
                 font=("Segoe UI", 8)).pack(side="left", padx=(0, 6))

        self._hist_filters = {}
        filter_cols = [
            ("fecha",    "Fecha",    10),
            ("codigo",   "Código",   8),
            ("nombre",   "Nombre",   14),
            ("tipo",     "Tipo",     7),
            ("cantidad", "Cantidad", 7),
            ("nota",     "Nota",     12),
        ]
        for col, placeholder, w in filter_cols:
            v = tk.StringVar()
            v.trace_add("write", lambda *_, c=col: self.refresh_historial())
            self._hist_filters[col] = v
            frm = tk.Frame(filter_frame, bg=BG)
            frm.pack(side="left", padx=3)
            tk.Label(frm, text=placeholder, bg=BG, fg=TEXT_DIM,
                     font=("Segoe UI", 7)).pack(anchor="w")
            entry(frm, textvariable=v, width=w).pack()

        cols_h = ("orden", "fecha", "codigo", "nombre", "tipo", "cantidad", "monto", "saldado", "nota")
        hdrs_h = ("Orden", "Fecha", "Código", "Nombre", "Tipo", "Cant.", "Monto", "✓", "Nota")

        tree_frame = tk.Frame(parent, bg=BG)
        tree_frame.pack(fill="both", expand=True, padx=10, pady=(0, 8))

        self.tree_hist = ttk.Treeview(tree_frame, columns=cols_h, show="headings", selectmode="browse")
        self.tree_hist.bind("<Button-1>", lambda e: "break" if self.tree_hist.identify_region(e.x, e.y) == "separator" else None)
        widths_h = (100, 110, 85, 180, 90, 55, 75, 30, 150)
        for col, hdr, w in zip(cols_h, hdrs_h, widths_h):
            self.tree_hist.heading(col, text=hdr)
            self.tree_hist.column(col, width=w, anchor="w" if col in ("nombre","nota") else "center")
        self.tree_hist.column("saldado", width=30, minwidth=30, anchor="center")

        self.tree_hist.tag_configure("entrada",          foreground=SUCCESS)
        self.tree_hist.tag_configure("salida",           foreground=DANGER)
        self.tree_hist.tag_configure("forzado",          foreground=WARNING, font=("Segoe UI", 9, "italic"))
        self.tree_hist.tag_configure("agrupado",         foreground=ACCENT2, font=("Segoe UI", 9, "bold"))
        self.tree_hist.tag_configure("detalle_agrupado", background=ROW_ALT)

        sb_h = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree_hist.yview)
        self.tree_hist.configure(yscrollcommand=sb_h.set)
        self.tree_hist.pack(side="left", fill="both", expand=True)
        sb_h.pack(side="right", fill="y")

        # Estado de filas expandidas en modo agrupado
        self._hist_expandidos = set()
        self._hist_movs_cache = {}
        self._hist_iid_map    = {}    # iid fila agrupada  -> producto_id
        self._hist_movid_map  = {}    # iid fila individual -> mov_id
        self.tree_hist.bind("<Double-1>", self._toggle_hist_grupo)
        self.tree_hist.bind("<Button-1>", self._on_hist_click)
        # Atajos helper text (sutil en el fondo)
        lbl_atajos_hist = tk.Label(parent, text="Atajos: [↑/↓] Navegar  •  [Enter] Expandir Grupo / Ver Detalle", bg=BG, fg=TEXT_DIM, font=("Segoe UI", 8))
        lbl_atajos_hist.pack(side="bottom", anchor="w", padx=14, pady=(0, 6))

        def is_hist_tab():
            try: return self.notebook.index(self.notebook.select()) == 1
            except: return False

        def _hist_select_first(event):
            if not is_hist_tab(): return
            sel = self.tree_hist.selection()
            if not sel:
                items = self.tree_hist.get_children()
                if items:
                    self.tree_hist.selection_set(items[0])
                    self.tree_hist.focus(items[0])
                    self.tree_hist.focus_set()
                    return "break"
            else:
                if not isinstance(event.widget, ttk.Treeview):
                    self.tree_hist.focus_set()
                    return "break"
        self.bind("<Down>", _hist_select_first, add="+")
        self.bind("<Up>", _hist_select_first, add="+")

        def _on_hist_enter(event):
            if not is_hist_tab(): return
            sel = self.tree_hist.selection()
            if not sel: return
            self._toggle_hist_grupo(event)
            return "break"
            
        self.bind("<Return>", _on_hist_enter, add="+")

    # ── Pestaña Alertas ───────────────────────────

    def _build_tab_alertas(self, parent):
        tk.Frame(parent, bg=BG, height=8).pack()

        self.lbl_alertas_titulo = tk.Label(
            parent, text="", bg=BG, fg=WARNING,
            font=("Segoe UI", 12, "bold"))
        self.lbl_alertas_titulo.pack(pady=(8, 4))

        cols_a = ("codigo", "nombre", "categoria", "stock", "minimo")
        hdrs_a = ("Código", "Nombre", "Categoría", "Stock actual", "Stock mínimo")

        tree_frame = tk.Frame(parent, bg=BG)
        tree_frame.pack(fill="both", expand=True, padx=10, pady=(0, 8))

        self.tree_alertas = ttk.Treeview(tree_frame, columns=cols_a, show="headings", selectmode="browse")
        self.tree_alertas.bind("<Button-1>", lambda e: "break" if self.tree_alertas.identify_region(e.x, e.y) == "separator" else None)
        widths_a = (100, 300, 160, 120, 120)
        for col, hdr, w in zip(cols_a, hdrs_a, widths_a):
            self.tree_alertas.heading(col, text=hdr)
            self.tree_alertas.column(col, width=w, anchor="w" if col in ("nombre","categoria") else "center")

        self.tree_alertas.tag_configure("bajo",     foreground=WARNING)
        self.tree_alertas.tag_configure("cero",     foreground=DANGER)

        sb_a = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree_alertas.yview)
        self.tree_alertas.configure(yscrollcommand=sb_a.set)
        self.tree_alertas.pack(side="left", fill="both", expand=True)
        sb_a.pack(side="right", fill="y")

        styled_btn(parent, "🔄 Actualizar alertas", self.refresh_alertas,
                   color=BG3).pack(pady=8)

    # ── Pestaña Edición Masiva ───────────────────────────



    # ── Refresh de datos ─────────────────────────

    def refresh_productos(self):
        filtro = self.v_buscar.get().strip() if hasattr(self, "v_buscar") else ""
        if filtro == "Buscar (F1)...": 
            filtro = ""
            
        cat    = self.v_cat_filtro.get().strip() if hasattr(self, "v_cat_filtro") else ""
        marca  = self.v_marca_filtro.get().strip() if hasattr(self, "v_marca_filtro") else ""
        
        productos_todos = db.get_productos(filtro)

        if cat and cat != "Cualquiera":
            productos_todos = [p for p in productos_todos if p["categoria"] == cat]
        if marca and marca != "Cualquiera":
            productos_todos = [p for p in productos_todos if p.get("marca", "") == marca]

        # Separar seleccionados de no seleccionados
        seleccionados = [p for p in productos_todos if str(p["id"]) in self._seleccionados_pos]
        no_sel = [p for p in productos_todos if str(p["id"]) not in self._seleccionados_pos]
        
        # Y recuperar aquellos que están seleccionados pero no figuran por filtro
        sel_ids = {str(p["id"]) for p in seleccionados}
        for sid in self._seleccionados_pos:
            if sid not in sel_ids:
                p_db = db.get_producto_by_id(int(sid))
                if p_db:
                    seleccionados.append(p_db)
                    
        productos_final = seleccionados + no_sel

        try:
            y_scroll = self.tree.yview()
        except tk.TclError:
            y_scroll = None

        self.tree.delete(*self.tree.get_children())

        for i, p in enumerate(productos_final):
            precio_str = f"${p['precio']:,.2f}"
            tags = []
            
            is_sel = str(p["id"]) in self._seleccionados_pos
            sel_str = "[F2]" if is_sel else "[ ]"
            
            if is_sel:
                tags.append("seleccionado")
            elif p["stock"] == 0:
                tags.append("sin_stock")
            elif p["stock"] < p["minimo"]:
                 tags.append("stock_bajo" if i % 2 == 0 else "alt_bajo")
            elif p.get("nota", "").strip():
                 tags.append("con_nota")
            elif i % 2 == 1:
                 tags.append("alt")

            self.tree.insert("", "end", iid=str(p["id"]),
                             values=(sel_str, p["codigo"], p["nombre"], p["categoria"], p["marca"],
                                     fmt_qty(p["stock"], p.get("por_peso")), fmt_qty(p["minimo"], p.get("por_peso")), precio_str),
                             tags=tuple(tags))

        total = len(productos_final)
        self.lbl_total.config(text=f"{total} producto{'s' if total != 1 else ''} ({len(self._seleccionados_pos)} seleccionados)")
        self.refresh_alertas()
        
        if y_scroll:
            try:
                self.tree.yview_moveto(y_scroll[0])
                self.tree.after(10, lambda: self.tree.yview_moveto(y_scroll[0]))
            except tk.TclError:
                pass

        if hasattr(self, "btn_carrito"):
            count = len(self._seleccionados_pos)
            if count > 0:
                self.btn_carrito.config(text=f"🛒 Carrito (F3) ({count})")
            else:
                self.btn_carrito.config(text="🛒 Carrito (F3)")

        # Refrescar listado categorías (por si hay nuevas)
        if hasattr(self, "cmb_cat"):
            cats = ["Cualquiera"] + db.get_categorias()
            if list(self.cmb_cat["values"]) != cats:
                self.cmb_cat.config(values=cats)
                
        # Refrescar listado marcas (por si hay nuevas)
        if hasattr(self, "cmb_marca"):
            marcas = ["Cualquiera"] + db.get_marcas()
            if list(self.cmb_marca["values"]) != marcas:
                self.cmb_marca.config(values=marcas)

    def _start_drag_sel_prod(self, event):
        tree = event.widget
        region = tree.identify_region(event.x, event.y)
        if region == "separator":
            return "break"
        if region == "cell":
            col = tree.identify_column(event.x)
            if col == "#1":
                iid = tree.identify_row(event.y)
                if iid:
                    self._drag_start_iid = iid
                    if tree == self.tree:
                        self._drag_start_action = iid not in self._seleccionados_pos
                    elif getattr(self, "tree_ed", None) and tree == self.tree_ed:
                        self._drag_start_action = iid not in self._seleccionados

    def _drag_sel_prod(self, event):
        tree = event.widget
        if not getattr(self, "_drag_start_iid", None): return
        region = tree.identify_region(event.x, event.y)
        if region == "cell":
            col = tree.identify_column(event.x)
            if col == "#1":
                iid = tree.identify_row(event.y)
                if iid:
                    changed = False
                    if tree == self.tree:
                        if getattr(self, "_drag_start_action", False) and iid not in self._seleccionados_pos:
                            self._seleccionados_pos.add(iid)
                            changed = True
                        elif not getattr(self, "_drag_start_action", False) and iid in self._seleccionados_pos:
                            self._seleccionados_pos.remove(iid)
                            if hasattr(self, "_cantidades_carrito"):
                                self._cantidades_carrito.pop(iid, None)
                            changed = True
                        if changed:
                            self.refresh_productos()
                    elif getattr(self, "tree_ed", None) and tree == self.tree_ed:
                        if getattr(self, "_drag_start_action", False) and iid not in self._seleccionados:
                            self._seleccionados.add(iid)
                            changed = True
                        elif not getattr(self, "_drag_start_action", False) and iid in self._seleccionados:
                            self._seleccionados.remove(iid)
                            changed = True
                        if changed:
                            if hasattr(self, 'refresh_tab_edicion'):
                                self.refresh_tab_edicion()


    def _toggle_sel_prod(self, event):
        tree = event.widget
        self._drag_start_iid = None # Reset drag
        region = tree.identify_region(event.x, event.y)
        if region == "cell":
            iid = tree.identify_row(event.y)
            if iid:
                if tree == self.tree:
                    self._last_selected_iid = iid
                elif getattr(self, "tree_ed", None) and tree == self.tree_ed:
                    self._last_selected_iid_ed = iid

                col = tree.identify_column(event.x)
                if col == "#1": # checkbox column
                    if tree == self.tree:
                        if iid in self._seleccionados_pos:
                            self._seleccionados_pos.remove(iid)
                            if hasattr(self, "_cantidades_carrito"):
                                self._cantidades_carrito.pop(iid, None)
                        else:
                            self._seleccionados_pos.add(iid)
                        self.refresh_productos()
                        try:
                            tree.focus_set()
                            tree.focus(iid)
                            tree.selection_set(iid)
                        except Exception as e: logging.exception("Error silencioso capturado:")
                    elif getattr(self, "tree_ed", None) and tree == self.tree_ed:
                        if iid in self._seleccionados:
                            self._seleccionados.remove(iid)
                        else:
                            self._seleccionados.add(iid)
                        # Si estamos en modo stock rápido no recargamos el tree (rompe la edición)
                        if not getattr(self, '_modo_stock_rapido', False):
                            if hasattr(self, 'refresh_tab_edicion'):
                                self.refresh_tab_edicion()
                            try:
                                tree.focus_set()
                                tree.focus(iid)
                                tree.selection_set(iid)
                            except Exception as e: logging.exception("Error silencioso capturado:")


    def _shift_toggle_sel_prod(self, event):
        tree = event.widget
        region = tree.identify_region(event.x, event.y)
        if region != "cell": return
        
        current_iid = tree.identify_row(event.y)
        
        is_pos_tab = (tree == self.tree)
        is_ed_tab = (getattr(self, "tree_ed", None) and tree == self.tree_ed)

        if is_pos_tab:
            last_sel = getattr(self, "_last_selected_iid", None)
            sel_set = self._seleccionados_pos
        elif is_ed_tab:
            last_sel = getattr(self, "_last_selected_iid_ed", None)
            sel_set = self._seleccionados
        else:
            return

        if not current_iid or not last_sel:
            return self._toggle_sel_prod(event)
            
        items = tree.get_children()
        try:
            start_idx = items.index(last_sel)
            end_idx = items.index(current_iid)
        except ValueError:
            return
            
        if start_idx > end_idx:
            start_idx, end_idx = end_idx, start_idx
            
        target_state = current_iid not in sel_set

        for i in range(start_idx, end_idx + 1):
            iid = items[i]
            if target_state:
                sel_set.add(iid)
            else:
                sel_set.discard(iid)
                if is_pos_tab and hasattr(self, "_cantidades_carrito"):
                    self._cantidades_carrito.pop(iid, None)
            
        if is_pos_tab:
            self._last_selected_iid = current_iid
            self.refresh_productos()
        elif is_ed_tab:
            self._last_selected_iid_ed = current_iid
            if hasattr(self, 'refresh_tab_edicion'):
                self.refresh_tab_edicion()

    def _deseleccionar_todos(self):
        self._seleccionados_pos.clear()
        if hasattr(self, "_cantidades_carrito"):
            self._cantidades_carrito.clear()
        self.refresh_productos()

    def _deseleccionar_todos_edicion(self):
        self._seleccionados.clear()
        if hasattr(self, 'refresh_tab_edicion'):
            self.refresh_tab_edicion()
        
    def _seleccionar_todos_visibles(self):
        for item in self.tree.get_children():
            self._seleccionados_pos.add(item)
        self.refresh_productos()


    def refresh_historial(self):
        movs = db.get_movimientos(limit=500)
        self.tree_hist.delete(*self.tree_hist.get_children())
        self._hist_movid_map = {}

        # Aplicar filtros por columna
        if hasattr(self, "_hist_filters"):
            col_keys = ["fecha", "codigo", "nombre", "tipo", "cantidad", "nota"]
            for col in col_keys:
                filtro = self._hist_filters[col].get().strip().lower()
                if not filtro:
                    continue
                if col == "tipo":
                    movs = [m for m in movs
                            if filtro in ("▲ entrada" if m["tipo"] == "entrada" else "▼ salida").lower()
                            or filtro in m["tipo"].lower()]
                elif col == "cantidad":
                    movs = [m for m in movs if filtro in str(m.get("cantidad", ""))]
                else:
                    movs = [m for m in movs if filtro in str(m.get(col, "")).lower()]

        agrupar_modo = self.v_agrupar.get() if hasattr(self, "v_agrupar") else "Sin agrupar"

        if agrupar_modo == "Producto":
            from collections import defaultdict
            grupos = defaultdict(lambda: {"entradas": 0, "salidas": 0, "codigo": "", "nombre": "", "fechas": [], "movs": [], "monto_total": 0.0})
            for m in movs:
                key = m["producto_id"]
                grupos[key]["codigo"] = m["codigo"]
                grupos[key]["nombre"] = m["nombre"]
                grupos[key]["fechas"].append(m["fecha"])
                grupos[key]["movs"].append(m)
                if m["tipo"] == "entrada":
                    grupos[key]["entradas"] += m["cantidad"]
                else:
                    grupos[key]["salidas"] += m["cantidad"]
                    if m.get("precio") is not None:
                        grupos[key]["monto_total"] += float(m["precio"])

            self._hist_movs_cache = {k: g["movs"] for k, g in grupos.items()}
            self._hist_iid_map    = {}

            for key, g in grupos.items():
                fecha_rango = g["fechas"][-1][:10] if g["fechas"] else ""
                expandido   = key in self._hist_expandidos
                icono       = "▼" if expandido else "▶"
                tipo_txt    = f"▲{fmt_qty(g['entradas'])} / ▼{fmt_qty(g['salidas'])}"
                monto_txt   = f"${g['monto_total']:,.2f}" if g['monto_total'] > 0 else ""

                iid = self.tree_hist.insert("", "end",
                                            values=("", fecha_rango, g["codigo"], f"{icono} {g['nombre']}",
                                                    tipo_txt, "", monto_txt, "", ""),
                                            tags=("agrupado",))
                self._hist_iid_map[iid] = key

                if expandido:
                    for m in g["movs"]:
                        t_txt = "  ▲ Entrada" if m["tipo"] == "entrada" else "  ▼ Salida"
                        sal = "☑" if m.get("saldado") else ("☐" if m.get("forzado") else "")
                        tags_row = ["detalle_agrupado"]
                        if m.get("forzado"): tags_row.append("forzado")
                        elif m["tipo"] == "entrada": tags_row.append("entrada")
                        else: tags_row.append("salida")
                        precio_val = m.get("precio")
                        monto_txt_m = f"${precio_val:,.2f}" if precio_val is not None else ""
                        ordem = m.get("grupo_id") or ""
                        mid = self.tree_hist.insert("", "end",
                                                    values=(ordem, m["fecha"], "", "",
                                                            t_txt, fmt_qty(m["cantidad"]), monto_txt_m, sal, m.get("nota", "")),
                                                    tags=tuple(tags_row))
                        self._hist_movid_map[mid] = m["id"]

        elif agrupar_modo == "Orden de Venta":
            from collections import defaultdict
            # Orden preservado: más nuevo primero (movs ya viene DESC)
            orden_keys = []  # para mantener el orden de aparición
            grupos = {}
            for m in movs:
                key = m.get("grupo_id") or f"#sin-orden/{m['id']}"
                if key not in grupos:
                    grupos[key] = {"fecha": m["fecha"], "movs": [], "monto_total": 0.0,
                                   "n_entradas": 0, "n_salidas": 0}
                    orden_keys.append(key)
                grupos[key]["movs"].append(m)
                if m["tipo"] == "entrada":
                    grupos[key]["n_entradas"] += 1
                else:
                    grupos[key]["n_salidas"] += 1
                if m.get("precio") is not None:
                    grupos[key]["monto_total"] += float(m["precio"])

            self._hist_movs_cache = grupos
            self._hist_iid_map    = {}

            for key in orden_keys:
                g = grupos[key]
                expandido = key in self._hist_expandidos
                icono     = "▼" if expandido else "▶"
                n_items   = len(g["movs"])
                partes = []
                if g["n_entradas"]: partes.append(f"▲{g['n_entradas']}")
                if g["n_salidas"]:  partes.append(f"▼{g['n_salidas']}")
                tipo_txt  = " / ".join(partes) if partes else "-"
                monto_txt = f"${g['monto_total']:,.2f}" if g["monto_total"] > 0 else ""
                orden_lbl = key if not key.startswith("#sin-orden") else f"(individual #{m['id']})"

                iid = self.tree_hist.insert("", "end",
                                            values=(f"{icono} {orden_lbl}", g["fecha"][:16], "", "",
                                                    tipo_txt, f"{n_items} item{'s' if n_items != 1 else ''}",
                                                    monto_txt, "", ""),
                                            tags=("agrupado",))
                self._hist_iid_map[iid] = key

                if expandido:
                    for m in g["movs"]:
                        t_txt = "  ▲ Entrada" if m["tipo"] == "entrada" else "  ▼ Salida"
                        sal = "☑" if m.get("saldado") else ("☐" if m.get("forzado") else "")
                        tags_row = ["detalle_agrupado"]
                        if m.get("forzado"): tags_row.append("forzado")
                        elif m["tipo"] == "entrada": tags_row.append("entrada")
                        else: tags_row.append("salida")
                        precio_val = m.get("precio")
                        monto_txt_m = f"${precio_val:,.2f}" if precio_val is not None else ""
                        mid = self.tree_hist.insert("", "end",
                                                    values=("", m["fecha"], m["codigo"],
                                                            f"      {m['nombre']}",
                                                            t_txt, fmt_qty(m["cantidad"]),
                                                            monto_txt_m, sal, m.get("nota", "")),
                                                    tags=tuple(tags_row))
                        self._hist_movid_map[mid] = m["id"]

        else:  # "Sin agrupar"
            for m in movs:
                tipo_txt  = "▲ Entrada" if m["tipo"] == "entrada" else "▼ Salida"
                # saldado solo se muestra si el movimiento tuvo forzado/excepción
                sal = "☑" if m.get("saldado") else ("☐" if m.get("forzado") else "")
                tags_row = []
                if m.get("forzado"): tags_row.append("forzado")
                else: tags_row.append(m["tipo"])
                precio_val = m.get("precio")
                monto_txt  = f"${precio_val:,.2f}" if precio_val is not None else ""
                orden_label = m.get("grupo_id") or ""
                iid = self.tree_hist.insert("", "end",
                                            values=(orden_label, m["fecha"], m["codigo"], m["nombre"],
                                                    tipo_txt, fmt_qty(m["cantidad"]), monto_txt, sal, m.get("nota", "")),
                                            tags=tuple(tags_row))
                self._hist_movid_map[iid] = m["id"]

    def _toggle_hist_grupo(self, event=None):
        """Doble-click o Enter: expande/colapsa agrupados, abre detalle en filas individuales."""
        iid = None
        if event and hasattr(event, "y") and event.y > 0 and getattr(event, "keysym", "") != "Return":
            iid = self.tree_hist.identify_row(event.y)
        if not iid:
            sel = self.tree_hist.selection()
            if sel: iid = sel[0]
            else: return
            
        agrupar_modo = self.v_agrupar.get() if hasattr(self, "v_agrupar") else "Sin agrupar"

        if iid in self._hist_iid_map and agrupar_modo != "Sin agrupar":
            # Es una fila de cabecera agrupada → toggle expand
            key = self._hist_iid_map[iid]
            if key in self._hist_expandidos:
                self._hist_expandidos.discard(key)
            else:
                self._hist_expandidos.add(key)
            self.refresh_historial()
            
            # Restaurar el foco y la selección a la fila que acabamos de clickear
            for new_iid, k in self._hist_iid_map.items():
                if k == key:
                    self.tree_hist.selection_set(new_iid)
                    self.tree_hist.focus(new_iid)
                    self.tree_hist.see(new_iid)
                    break
        elif iid in self._hist_movid_map:
            # Es una fila de movimiento individual → abrir detalle
            mov_id = self._hist_movid_map[iid]
            self._abrir_detalle_mov(mov_id)

    def _abrir_detalle_mov(self, mov_id):
        """Muestra un popup con el detalle completo del movimiento."""
        # Buscar el movimiento en caché
        movs = db.get_movimientos(limit=2000)
        m = next((x for x in movs if x["id"] == mov_id), None)
        if not m:
            return

        dlg = tk.Toplevel(self)
        dlg.title("Detalle de Movimiento")
        dlg.configure(bg=BG)
        dlg.resizable(False, False)
        dlg.grab_set()

        tipo_txt = "▲ Entrada" if m["tipo"] == "entrada" else "▼ Salida"
        color_tipo = SUCCESS if m["tipo"] == "entrada" else DANGER
        if m.get("forzado"): color_tipo = WARNING

        def row(label, value, colorfg=TEXT):
            frm = tk.Frame(dlg, bg=BG)
            frm.pack(fill="x", padx=20, pady=3)
            tk.Label(frm, text=label, bg=BG, fg=TEXT_DIM,
                     font=("Segoe UI", 9), width=14, anchor="w").pack(side="left")
            tk.Label(frm, text=str(value), bg=BG, fg=colorfg,
                     font=("Segoe UI", 9, "bold"), anchor="w", wraplength=320,
                     justify="left").pack(side="left", fill="x", expand=True)

        tk.Frame(dlg, bg=BG, height=8).pack()
        tk.Label(dlg, text=tipo_txt, bg=BG, fg=color_tipo,
                 font=("Segoe UI", 13, "bold")).pack(padx=20, anchor="w")
        tk.Frame(dlg, bg=BG2, height=1).pack(fill="x", padx=16, pady=6)

        row("Nº Orden:",   m.get("grupo_id") or "-")
        row("Fecha:",      m["fecha"])
        row("Código:",     m.get("codigo", "-"))
        row("Producto:",   m.get("nombre", "-"))
        row("Cantidad:",   fmt_qty(m["cantidad"]))
        precio_val = m.get("precio")
        row("Monto:",      f"${precio_val:,.2f}" if precio_val is not None else "-")
        if m.get("forzado"):
            row("⚠ Tipo:",  "Forzado (stock insuficiente)", WARNING)
        if m.get("saldado"):
            row("☑ Saldado:", "Sí", SUCCESS)

        nota = m.get("nota", "").strip()
        tk.Frame(dlg, bg=BG2, height=1).pack(fill="x", padx=16, pady=6)
        tk.Label(dlg, text="Nota:", bg=BG, fg=TEXT_DIM,
                 font=("Segoe UI", 9), anchor="w").pack(padx=20, anchor="w")
        nota_box = tk.Text(dlg, bg=BG3, fg=TEXT, font=("Segoe UI", 9),
                           width=42, height=4, relief="flat", wrap="word",
                           padx=8, pady=6, state="normal")
        nota_box.insert("1.0", nota if nota else "(sin nota)")
        nota_box.config(state="disabled")
        nota_box.pack(padx=20, pady=(0, 4), fill="x")

        tk.Frame(dlg, bg=BG, height=8).pack()
        tk.Button(dlg, text="Cerrar", command=dlg.destroy,
                  bg=BG3, fg=TEXT, relief="flat",
                  padx=16, pady=6, font=("Segoe UI", 9, "bold"),
                  activebackground=ACCENT, cursor="hand2").pack(pady=(0, 14))

        dlg.bind("<Return>", lambda e: dlg.destroy())
        dlg.bind("<Escape>", lambda e: dlg.destroy())
        dlg.update_idletasks()
        # Centrar
        w, h = dlg.winfo_width(), dlg.winfo_height()
        x = self.winfo_x() + (self.winfo_width()  - w) // 2
        y = self.winfo_y() + (self.winfo_height() - h) // 2
        dlg.geometry(f"+{x}+{y}")

    def _exportar_ventas(self):
        """Exporta todos los movimientos a Excel (o CSV) y ofrece vaciar el historial."""
        from tkinter import filedialog
        
        movs = db.get_movimientos(limit=999999)
        if not movs:
            messagebox.showinfo("Sin datos", "El historial está vacío, no hay nada que exportar.", parent=self)
            return

        # Nombres de exportación basados en fechas
        fechas = [m["fecha"][:10] for m in movs if "fecha" in m and m["fecha"]]
        if fechas:
            fechas_ordenadas = sorted(fechas)
            primera = fechas_ordenadas[0].replace("-", "")
            ultima = fechas_ordenadas[-1].replace("-", "")
            if primera == ultima:
                nombre_default = f"Ventas_{primera}"
            else:
                nombre_default = f"Ventas_{primera}_{ultima}"
        else:
            import datetime
            fecha_str = datetime.datetime.now().strftime("%Y%m%d_%H%M")
            nombre_default = f"Ventas_{fecha_str}"

        # Intentar exportar a Excel primero
        try:
            import openpyxl  # noqa: F401
            excel_disponible = True
        except ImportError:
            excel_disponible = False

        if excel_disponible:
            filetypes = [("Excel", "*.xlsx"), ("CSV", "*.csv"), ("Todos", "*.*")]
        else:
            filetypes = [("CSV", "*.csv"), ("Todos", "*.*")]

        filepath = filedialog.asksaveasfilename(
            parent=self,
            title="Exportar historial de ventas",
            initialfile=nombre_default,
            defaultextension=".xlsx" if excel_disponible else ".csv",
            filetypes=filetypes,
        )
        if not filepath:
            return  # Cancelado

        try:
            if filepath.lower().endswith(".xlsx") and excel_disponible:
                import openpyxl
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Historial de ventas"
                headers = ["ID", "Nº Orden", "Fecha", "Código", "Producto",
                           "Tipo", "Cantidad", "Monto", "Forzado", "Saldado", "Nota"]
                ws.append(headers)
                # Estilo de encabezado
                from openpyxl.styles import Font, PatternFill
                for cell in ws[1]:
                    cell.font  = Font(bold=True, color="FFFFFF")
                    cell.fill  = PatternFill("solid", fgColor="1a3a2a")
                for m in movs:
                    ws.append([
                        m["id"],
                        m.get("grupo_id") or "",
                        m["fecha"],
                        m.get("codigo", ""),
                        m.get("nombre", ""),
                        m["tipo"],
                        float(m["cantidad"]),
                        float(m["precio"]) if m.get("precio") is not None else "",
                        "Sí" if m.get("forzado") else "No",
                        "Sí" if m.get("saldado") else "No",
                        m.get("nota", "") or "",
                    ])
                # Ajustar ancho de columnas
                col_widths = [8, 20, 20, 14, 30, 10, 10, 12, 8, 8, 30]
                for i, w in enumerate(col_widths, 1):
                    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
                wb.save(filepath)
            else:
                import csv as csv_mod
                with open(filepath, "w", newline="", encoding="utf-8") as f:
                    writer = csv_mod.writer(f)
                    writer.writerow(["ID", "Nº Orden", "Fecha", "Código", "Producto",
                                     "Tipo", "Cantidad", "Monto", "Forzado", "Saldado", "Nota"])
                    for m in movs:
                        writer.writerow([
                            m["id"],
                            m.get("grupo_id") or "",
                            m["fecha"],
                            m.get("codigo", ""),
                            m.get("nombre", ""),
                            m["tipo"],
                            m["cantidad"],
                            m["precio"] if m.get("precio") is not None else "",
                            "Sí" if m.get("forzado") else "No",
                            "Sí" if m.get("saldado") else "No",
                            m.get("nota", "") or "",
                        ])
        except Exception as exc:
            messagebox.showerror("Error al exportar", str(exc), parent=self)
            return

        # Éxito: ofrecer vaciar historial
        resp = messagebox.askyesno(
            "Exportación completada",
            f"✅ Se exportaron {len(movs)} registros a:\n{filepath}\n\n"
            "¿Deseas vaciar el historial de ventas ahora?",
            parent=self
        )
        if resp:
            n = db.vaciar_movimientos()
            self.refresh_historial()
            messagebox.showinfo(
                "Historial vaciado",
                f"Se eliminaron {n} movimientos del historial.",
                parent=self
            )

    def refresh_alertas(self):
        bajos = db.get_stock_bajo()
        self.tree_alertas.delete(*self.tree_alertas.get_children())
        
        # Filtrar si hay busqueda
        busqueda = ""
        if hasattr(self, 'v_buscar_alertas'):
            val = self.v_buscar_alertas.get().strip().lower()
            if val and val != "buscar...":
                busqueda = val

        for p in bajos:
            if busqueda:
                texto_buscable = f"{p.get('codigo','')} {p.get('nombre','')} {p.get('categoria','')} {p.get('marca','')}".lower()
                if busqueda not in texto_buscable:
                    continue
                    
            tag = "cero" if p["stock"] == 0 else "bajo"
            self.tree_alertas.insert("", "end",
                                     values=(p["codigo"], p["nombre"], p["categoria"],
                                             fmt_qty(p["stock"], p.get("por_peso")), fmt_qty(p["minimo"], p.get("por_peso"))),
                                     tags=(tag,))
        n = len(self.tree_alertas.get_children())
        titulo = f"⚠  {n} producto{'s' if n != 1 else ''} con stock bajo" if n else "✔  Todos los productos tienen stock suficiente"
        self.lbl_alertas_titulo.config(text=titulo, fg=WARNING if n else SUCCESS)

    # ── Toggle de edici\u00f3n ──────────────────────────

    def _toggle_edicion(self):
        habilitado = not self.v_edicion.get()
        self.v_edicion.set(habilitado)
        self._aplicar_estado_edicion()

    def _aplicar_estado_edicion(self):
        habilitado = self.v_edicion.get()
        # Botón del header
        if habilitado:
            self.btn_lock.config(text="\U0001f513  Edición habilitada", bg=SUCCESS)
        else:
            self.btn_lock.config(text="\U0001f512  Edición bloqueada", bg="#8b3a3a")
        # Deshabilitar / habilitar botones de edición
        state = "normal" if habilitado else "disabled"
        for btn in self._btns_edicion:
            btn.config(state=state)
        if hasattr(self, 'btn_export_ventas'):
            self.btn_export_ventas.config(state=state)

    def _on_tab_change(self, event):
        tab = event.widget.tab("current", "text").strip()
        if "Historial" in tab:
            self.refresh_historial()
        elif "Alerta" in tab:
            self.refresh_alertas()
        elif "Edición" in tab or "Edicion" in tab:
            # Bloquear acceso si edición no está habilitada
            if not self.v_edicion.get():
                # Volver a la pestaña Productos
                self.notebook.select(0)
                messagebox.showwarning(
                    "Acceso restringido",
                    "La pestaña Edición solo está disponible cuando la edición está habilitada.\n"
                    "Activa el botón 🔓 Edición en el encabezado.",
                    parent=self
                )
                return
            self.refresh_tab_edicion()
    def refresh_tab_edicion(self):
        """Refresca la lista de productos en la pestaña Edición aplicando filtros."""
        if not hasattr(self, "tree_ed"):
            return
            
        # Refrescar dropdowns
        categorias = ["Todas", "Sin categoría"] + [c for c in db.get_categorias() if c]
        if hasattr(self, 'cmb_cat_ed'):
            self.cmb_cat_ed.config(values=categorias)
            if self.v_cat_filtro_ed.get() not in categorias: self.v_cat_filtro_ed.set("Todas")
        marcas = ["Todas", "Sin marca"] + [m for m in db.get_marcas() if m]
        if hasattr(self, 'cmb_marca_ed'):
            self.cmb_marca_ed.config(values=marcas)
            if self.v_marca_filtro_ed.get() not in marcas: self.v_marca_filtro_ed.set("Todas")
            
        try:
            y_scroll_ed = self.tree_ed.yview()
        except tk.TclError:
            y_scroll_ed = None
            
        self.tree_ed.delete(*self.tree_ed.get_children())
        prods = db.get_productos()
        min_stock = int(db.get_config("min_stock", 5))
        
        cat_filtrada = self.v_cat_filtro_ed.get() if hasattr(self, 'v_cat_filtro_ed') else "Todas"
        marca_filtrada = self.v_marca_filtro_ed.get() if hasattr(self, 'v_marca_filtro_ed') else "Todas"
        busqueda = ""
        if hasattr(self, 'v_buscar_edicion'):
            val = self.v_buscar_edicion.get().strip().lower()
            if val and val not in ["buscar...", "buscar (f1)..."]:
                busqueda = val
        
        visible_index = 0
        for p in prods:
            if busqueda:
                texto = f"{p.get('codigo','')} {p.get('nombre','')} {p.get('categoria','')} {p.get('marca','')}".lower()
                if busqueda not in texto:
                    continue
            c = p.get("categoria", "")
            if cat_filtrada == "Sin categoría" and c: continue
            if cat_filtrada != "Todas" and cat_filtrada != "Sin categoría" and c != cat_filtrada: continue
            
            m = p.get("marca", "")
            if marca_filtrada == "Sin marca" and m: continue
            if marca_filtrada != "Todas" and marca_filtrada != "Sin marca" and m != marca_filtrada: continue

            stock_val = p.get("stock", 0)
            if getattr(self, '_modo_stock_rapido', False) and p["id"] in getattr(self, '_cambios_stock_tmp', {}):
                stock_val = self._cambios_stock_tmp[p["id"]]
                
            try:
                stock_num = float(stock_val)
            except (TypeError, ValueError):
                stock_num = 0.0
                
            is_sel = str(p["id"]) in getattr(self, "_seleccionados", set())
            sel_str = "[F2]" if is_sel else "[ ]"

            if is_sel:
                tag = "seleccionado"
            elif stock_num <= 0:
                tag = "cero"
            elif stock_num < min_stock:
                tag = "bajo"
            elif visible_index % 2 == 1:
                tag = "alt"
            else:
                tag = ""
                
            self.tree_ed.insert("", "end", iid=str(p["id"]), tags=(tag,),
                                values=(sel_str, p.get("codigo",""), p.get("nombre",""),
                                        p.get("marca","") or "", p.get("categoria","") or "",
                                        fmt_qty(stock_num, p.get("por_peso", 0)), 
                                        f"${p.get('precio_costo',0):,.2f}", f"${p.get('precio',0):,.2f}"))
            visible_index += 1
            
        if y_scroll_ed:
            try:
                self.tree_ed.yview_moveto(y_scroll_ed[0])
                self.tree_ed.after(10, lambda: self.tree_ed.yview_moveto(y_scroll_ed[0]))
            except tk.TclError:
                pass

    def _editar_producto_desde_tree_ed(self):
        """Abre el diálogo de edición del producto seleccionado en tree_ed."""
        self._editar_producto()


    # ── Acciones de productos ─────────────────────

    def _abrir_aumento(self):
        ids = self._get_selected_productos_ids()
        productos = [db.get_producto_by_id(pid) for pid in ids]
        dlg = AumentoMasivoDialog(self, productos_seleccionados=productos)
        if getattr(dlg, "resultado", False):
            self.refresh_productos()
            if hasattr(self, 'refresh_tab_edicion'):
                self.refresh_tab_edicion()

    def _redondear_up(self):
        ids = self._get_selected_productos_ids()
        if not ids:
            messagebox.showwarning("Atención", "Debe seleccionar al menos un producto para redondear.")
            return
        if messagebox.askyesno("Confirmar", f"¿Seguro que desea redondear hacia ARRIBA los {len(ids)} productos seleccionados?"):
            db.redondear_precios_masivo(ids=ids, direccion="arriba")
            messagebox.showinfo("Éxito", "Precios redondeados hacia arriba.")
            self.refresh_productos()
            if hasattr(self, 'refresh_tab_edicion'):
                self.refresh_tab_edicion()

    def _redondear_dn(self):
        ids = self._get_selected_productos_ids()
        if not ids:
            messagebox.showwarning("Atención", "Debe seleccionar al menos un producto para redondear.")
            return
        if messagebox.askyesno("Confirmar", f"¿Seguro que desea redondear hacia ABAJO los {len(ids)} productos seleccionados?"):
            db.redondear_precios_masivo(ids=ids, direccion="abajo")
            messagebox.showinfo("Éxito", "Precios redondeados hacia abajo.")
            self.refresh_productos()
            if hasattr(self, 'refresh_tab_edicion'):
                self.refresh_tab_edicion()



    def _get_selected_productos_ids(self):
        """Devuelve una lista de IDs de las selecciones explícitas, o fallback en Historial/Productos."""
        tab = self.notebook.tab("current", "text").strip()
        if "Edición" in tab or "Edicion" in tab:
            return [int(s) for s in getattr(self, "_seleccionados", set())]
        else:
            return [int(s) for s in self.tree.selection()]

    def _get_selected_producto(self):
        ids = self._get_selected_productos_ids()
        if not ids:
            messagebox.showinfo("Selección", "Seleccione un producto de la lista.")
            return None
        return db.get_producto_by_id(ids[0])

    def _nuevo_producto(self):
        dlg = ProductoDialog(self)
        if dlg.resultado:
            try:
                db.crear_producto(*dlg.resultado)
                self.refresh_productos()
            except Exception as e:
                messagebox.showerror("Error al crear", str(e))

    def _editar_producto(self):
        ids = self._get_selected_productos_ids()
        if not ids:
            messagebox.showinfo("Selección", "Seleccione al menos un producto de la lista.")
            return
            
        if len(ids) == 1:
            prod = db.get_producto_by_id(ids[0])
            if not prod:
                return
            dlg = ProductoDialog(self, prod)
            if dlg.resultado:
                try:
                    db.actualizar_producto(prod["id"], *dlg.resultado)
                    self.refresh_productos()
                except Exception as e:
                    messagebox.showerror("Error al actualizar", str(e))
        else:
            # Edición masiva
            dlg = EdicionMasivaDialog(self, count=len(ids))
            if dlg.resultado:
                try:
                    db.actualizar_productos_masivo(ids, dlg.resultado)
                    self.refresh_productos()
                    messagebox.showinfo("Edición Masiva", f"Se actualizaron {len(ids)} productos exitosamente.")
                except Exception as e:
                    messagebox.showerror("Error al actualizar", f"Ocurrió un error en la actualización masiva:\n{str(e)}")

    def _eliminar_producto(self):
        ids = self._get_selected_productos_ids()
        if not ids:
            messagebox.showinfo("Eliminar Producto", "Seleccione al menos un producto para eliminar.")
            return

        if len(ids) == 1:
            prod = db.get_producto_by_id(ids[0])
            if not prod: return
            msg = f"¿Eliminar «{prod['nombre']}»?\n\nSu historial se preservará pero no aparecerá en el inventario activo."
        else:
            msg = f"¿Eliminar {len(ids)} productos seleccionados?\n\nSu historial se preservará pero no aparecerán en el inventario activo."

        ok = messagebox.askyesno("Confirmar eliminación", msg, icon="warning")
        
        if ok:
            for pid in ids:
                db.eliminar_producto(pid)
            
            # Deseleccionar luego de borrar, para evitar ids fantasma
            self._deseleccionar_todos()
            self._deseleccionar_todos_edicion()
            self.refresh_productos()

    def _entrada_stock(self):
        prod = self._get_selected_producto()
        if not prod:
            return
        dlg = MovimientoDialog(self, prod, "entrada")
        if dlg.resultado:
            self.refresh_productos()

    def _abrir_carrito(self):
        if not hasattr(self, "_seleccionados_pos") or not self._seleccionados_pos:
            messagebox.showinfo("Carrito Vacío", "No hay productos seleccionados para iniciar una salida múltiple.")
            return

        ps_seleccionados = []
        cantidades = getattr(self, "_cantidades_carrito", {})
        for sid in list(self._seleccionados_pos):
            p_db = db.get_producto_by_id(int(sid))
            if p_db:
                p_db["_init_cant"] = cantidades.get(sid, 1)
                ps_seleccionados.append(p_db)
            
        dlg = SalidaMultipleDialog(self, ps_seleccionados)
        if getattr(dlg, "resultado", False):
            self._seleccionados_pos.clear()
            if hasattr(self, "_cantidades_carrito"):
                self._cantidades_carrito.clear()
            self.refresh_productos()

    def _abrir_reporte_capital(self):
        ReporteCapitalDialog(self)
        
    def _exportar_excel(self):
        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")],
            title="Guardar inventario como…"
        )
        if filepath:
            n = db.exportar_excel(filepath)
            messagebox.showinfo("Exportación exitosa", f"Se exportaron {n} productos a:\n{filepath}")

    def _importar_datos(self):
        dlg = ImportarDialog(self)
        if hasattr(dlg, "mapeo_resultado") and dlg.mapeo_resultado:
            # Procesar la importacion
            try:
                import importador
                # Ahora el importador lee por mapeo directamente, ya que contiene inicio/fin/columna
                filepath = dlg.filepath
                mapeo = dlg.mapeo_resultado
                items_list = importador.leer_datos(filepath, mapeo)

                if not items_list:
                    messagebox.showinfo("Sin datos", "No se encontraron datos para importar.", parent=self)
                    return

                nombres_nuevos = []
                nombres_actualizados = []

                for item in items_list:
                    nombre = str(item.get("nombre", "")).strip()
                    if not nombre:
                        continue

                    codigo = str(item.get("codigo", "")).strip()

                    categoria = str(item.get("categoria", "")).strip()
                    marca     = str(item.get("marca", "")).strip()
                    try: stock = int(item.get("stock", 0))
                    except: stock = 0
                    try: minimo = int(item.get("minimo", 0))
                    except: minimo = 0
                    try: precio = float(item.get("precio", 0.0))
                    except: precio = 0.0
                    try: precio_costo = float(item.get("precio_costo", 0.0))
                    except: precio_costo = 0.0


                    # Llamar al upsert
                    is_new, cambios = db.upsert_producto(codigo, nombre, categoria, marca, stock, minimo, precio, 0, precio_costo)
                    if is_new:
                        nombres_nuevos.append(nombre)
                    else:
                        if cambios:
                            nombres_actualizados.append((nombre, cambios))

                ImportacionResumenDialog(self, nombres_nuevos, nombres_actualizados)
                self.refresh_productos()
                
            except Exception as e:
                messagebox.showerror("Error de importación", f"Ocurrió un problema: {str(e)}")

    # ── Ordenamiento de columnas ──────────────────

    def _on_tree_double_click(self, event):
        """Alternar selección con doble clic."""
        region = self.tree.identify_region(event.x, event.y)
        if region == "cell":
            if not getattr(self, "v_edicion", None) or not self.v_edicion.get():
                iid = self.tree.identify_row(event.y)
                if iid:
                    if iid in self._seleccionados_pos:
                        self._seleccionados_pos.remove(iid)
                        if hasattr(self, "_cantidades_carrito"):
                            self._cantidades_carrito.pop(iid, None)
                    else:
                        self._seleccionados_pos.add(iid)
                    self.refresh_productos()
            else:
                self._editar_producto()

    def _on_tree_enter(self, event):
        """Atajo Enter en la tabla. Pide cantidad y añade al carrito si bloqueada, edita si desbloqueada."""
        sel = self.tree.selection()
        if not sel: return
        iid = sel[0]
        try: y_pos = self.tree.yview()
        except tk.TclError: y_pos = None

        if not getattr(self, "v_edicion", None) or not self.v_edicion.get():
            prod = db.get_producto_by_id(int(iid))
            if prod:
                if prod.get("por_peso", 0):
                    dlg = ModificarItemDialog(self, prod, cant_inicial=1.0, nota_inicial="")
                    if getattr(dlg, "resultado", False) and isinstance(dlg.resultado, dict):
                        cant = dlg.resultado["cant"]
                    else:
                        cant = None
                else:
                    from tkinter.simpledialog import askinteger
                    cant = askinteger("Añadir al Carrito", f"Cantidad a añadir de '{prod['nombre']}':", minvalue=1, initialvalue=1, parent=self)
                    
                if cant is not None:
                    if not hasattr(self, "_cantidades_carrito"):
                        self._cantidades_carrito = {}
                    self._cantidades_carrito[iid] = cant
                    self._seleccionados_pos.add(iid)
                    self.refresh_productos()
        else:
            self._editar_producto()
            
        self.tree.focus_set()
        self.tree.focus(iid)
        self.tree.selection_set(iid)
        if y_pos:
            self.tree.after(10, lambda: self.tree.yview_moveto(y_pos[0]))

    def _on_f6_pressed(self):
        """Atajo F6. Visualiza/edita la nota del producto."""
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("INFO", "Seleccione al menos un producto para gestionar su nota.", parent=self)
            return
            
        edicion_bloqueada = not getattr(self, "v_edicion", None) or not self.v_edicion.get()
        if edicion_bloqueada:
            # Si está bloqueada, solo ver la nota del primero
            prod = db.get_producto_by_id(int(sel[0]))
            nota = prod.get("nota", "") if prod else ""
            if nota:
                messagebox.showinfo(f"Nota para {prod.get('nombre')}", nota, parent=self)
            return

        # Edición habilitada: modificar
        if len(sel) == 1:
            prod = db.get_producto_by_id(int(sel[0]))
            nota_inicial = prod.get("nota", "") if prod else ""
            txt = simpledialog.askstring("Nota de Producto", f"Nota para '{prod.get('nombre')}':", initialvalue=nota_inicial, parent=self)
        else:
            txt = simpledialog.askstring("Nota Masiva", f"Se aplicará esta nota a {len(sel)} productos seleccionados:", parent=self)
            
        if txt is not None:
            txt = txt.strip()
            for iid in sel:
                db.actualizar_nota_producto(int(iid), txt)
            self.refresh_productos()
            self.tree.selection_set(sel)
            self.tree.focus(sel[0])


    def _on_hist_click(self, event):
        """Toggle saldado al hacer click en la columna '✓'."""
        col = self.tree_hist.identify_column(event.x)
        if col != "#8":  # columna saldado ('✓') es la 8va
            return
        iid = self.tree_hist.identify_row(event.y)
        if not iid or iid not in self._hist_movid_map:
            return
        mov_id = self._hist_movid_map[iid]
        db.toggle_saldado(mov_id)
        # Actualizar solo el valor en la fila sin refrescar todo
        vals = list(self.tree_hist.item(iid, "values"))
        vals[7] = "☑" if vals[7] == "☐" else "☐"
        self.tree_hist.item(iid, values=vals)



    def _abrir_configuracion(self):

        SettingsDialog(self)

        self.refresh_productos()



    def _sort_column(self, col):
        items = [(self.tree.set(k, col), k) for k in self.tree.get_children("")]
        rev = self._sort_col == col and not self._sort_rev
        
        if col == "sel":
            # Si estamos ordenando por la columna de selección F2, forzar '[F2]' arriba (o abajo si rev)
            items.sort(key=lambda t: 0 if t[0] == "[F2]" else 1, reverse=rev)
        else:
            try:
                items.sort(key=lambda t: float(t[0].replace("$","")) if t[0].replace("$","").replace(".","").isdigit() else t[0].lower(), reverse=rev)
            except Exception:
                items.sort(key=lambda t: t[0].lower(), reverse=rev)
                
        for idx, (_, k) in enumerate(items):
            self.tree.move(k, "", idx)
        self._sort_col = col
        self._sort_rev = rev


# ──────────────────────────────────────────────
#  Ventana de Configuracion
# ──────────────────────────────────────────────

class SettingsDialog(tk.Toplevel):
    def __init__(self, parent):
        super().__init__(parent)
        self.title("⚙  Configuración")
        self.configure(bg=BG2)
        self.bind("<Escape>", lambda e: self.destroy())
        self.resizable(False, False)
        self.grab_set()

        frame = tk.Frame(self, bg=BG2, padx=28, pady=24)
        frame.pack(fill="both", expand=True)

        tk.Label(frame, text="⚙  Configuración general",
                 bg=BG2, fg=ACCENT2, font=("Segoe UI", 13, "bold")).grid(
            row=0, column=0, columnspan=2, sticky="w", pady=(0, 18))

        sec = tk.LabelFrame(frame, text=" Stock mínimo por defecto ",
                            bg=BG2, fg=TEXT_DIM,
                            font=("Segoe UI", 9), bd=1, relief="groove")
        sec.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(0, 14), padx=2)

        tk.Label(sec, text=(
            "Aplica este valor a todos los productos\n"
            "cuyo stock mínimo sea 0 (no configurado)."
        ), bg=BG2, fg=TEXT_DIM, font=("Segoe UI", 8), justify="left").grid(
            row=0, column=0, columnspan=2, sticky="w", padx=10, pady=(6, 4))

        current = db.get_config("minimo_defecto", "0")
        self.v_minimo = tk.StringVar(value=current)

        tk.Label(sec, text="Valor:", bg=BG2, fg=TEXT,
                 font=("Segoe UI", 9)).grid(row=1, column=0, sticky="w", padx=10, pady=6)
        entry(sec, textvariable=self.v_minimo, width=10).grid(
            row=1, column=1, sticky="w", padx=(0, 10), pady=6)

        styled_btn(sec, "▶ Aplicar a productos con minimo=0",
                   self._aplicar_minimo, color=ACCENT).grid(
            row=2, column=0, columnspan=2, pady=(0, 10), padx=10)

        self.lbl_resultado = tk.Label(sec, text="", bg=BG2, fg=SUCCESS,
                                      font=("Segoe UI", 8, "italic"))
        self.lbl_resultado.grid(row=3, column=0, columnspan=2, pady=(0, 8))
        
        # Categorías y marcas personalizadas
        sec_custom = tk.LabelFrame(frame, text=" Categorías y Marcas Personalizadas ",
                            bg=BG2, fg=TEXT_DIM,
                            font=("Segoe UI", 9), bd=1, relief="groove")
        sec_custom.grid(row=2, column=0, columnspan=2, sticky="ew", pady=(0, 14), padx=2)
        
        tk.Label(sec_custom, text=(
            "Agregue categorías y marcas separadas por '|' (tubo)."
        ), bg=BG2, fg=TEXT_DIM, font=("Segoe UI", 8), justify="left").grid(
            row=0, column=0, columnspan=2, sticky="w", padx=10, pady=(6, 4))
            
        current_cats = db.get_config("custom_categorias", "")
        self.v_custom_cat = tk.StringVar(value=current_cats)
        
        tk.Label(sec_custom, text="Categorías:", bg=BG2, fg=TEXT,
                 font=("Segoe UI", 9)).grid(row=1, column=0, sticky="w", padx=10, pady=6)
        entry(sec_custom, textvariable=self.v_custom_cat, width=40).grid(
            row=1, column=1, sticky="w", padx=(0, 10), pady=6)
            
        current_marcas = db.get_config("custom_marcas", "")
        self.v_custom_marca = tk.StringVar(value=current_marcas)
        
        tk.Label(sec_custom, text="Marcas:", bg=BG2, fg=TEXT,
                 font=("Segoe UI", 9)).grid(row=2, column=0, sticky="w", padx=10, pady=6)
        entry(sec_custom, textvariable=self.v_custom_marca, width=40).grid(
            row=2, column=1, sticky="w", padx=(0, 10), pady=6)
            
        styled_btn(sec_custom, "▶ Guardar Listas",
                   self._guardar_listas, color=ACCENT).grid(
            row=3, column=0, columnspan=2, pady=(0, 10), padx=10)
            
        self.lbl_resultado_listas = tk.Label(sec_custom, text="", bg=BG2, fg=SUCCESS,
                                      font=("Segoe UI", 8, "italic"))
        self.lbl_resultado_listas.grid(row=4, column=0, columnspan=2, pady=(0, 8))

        styled_btn(frame, "✔ Cerrar", self.destroy, color=BG3).grid(
            row=3, column=0, columnspan=2, pady=(10, 0))

        self.wait_window()

    def _guardar_listas(self):
        cat_val = self.v_custom_cat.get().strip()
        marca_val = self.v_custom_marca.get().strip()
        db.set_config("custom_categorias", cat_val)
        db.set_config("custom_marcas", marca_val)
        self.lbl_resultado_listas.config(text="✓ Listas guardadas correctamente")
        self.after(2500, lambda: self.lbl_resultado_listas.config(text=""))

    def _aplicar_minimo(self):
        try:
            val = int(self.v_minimo.get())
            if val < 0:
                raise ValueError
        except ValueError:
            messagebox.showerror('Valor invalido', 'Ingrese un numero entero >= 0.', parent=self)
            return
        db.set_config('minimo_defecto', val)
        n = db.aplicar_minimo_defecto(val)
        self.lbl_resultado.config(
            text=f'✔  {n} producto(s) actualizados.' if n else 'Sin productos para actualizar.'
        )


if __name__ == '__main__':
    db.init_db()
    root = StockApp()
    root.mainloop()
